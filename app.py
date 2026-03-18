"""
ForecastWell Dashboard - Flask Application
Weather-based demand forecasting for HVAC/Consumer Durables
Enhanced per ForecastWell Guide - Night Temperature Priority!
"""
from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for, Response
from functools import wraps
from flask_cors import CORS
from datetime import datetime, timedelta
import os
import random
import math
import time
import threading
import queue
import json as json_module
import gzip
import io
import requests
from concurrent.futures import ThreadPoolExecutor
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from config import Config
from utils.weather_service import WeatherService
from utils.alert_engine import AlertEngine
from utils.data_processor import DataProcessor
from utils.supabase_client import SupabaseHandler
from utils.file_cache import FileCache

app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# Unique token generated fresh on every app start.
# Any session cookie from a previous run won't carry this token,
# so the user is forced to log in again after the app is restarted.
BOOT_TOKEN = os.urandom(16).hex()

# Performance timing middleware
@app.before_request
def before_request_timing():
    """Store request start time for performance monitoring"""
    request.start_time = time.time()

@app.after_request
def after_request_timing(response):
    """Add performance timing header"""
    if hasattr(request, 'start_time'):
        elapsed = time.time() - request.start_time
        response.headers['X-Response-Time'] = f'{elapsed*1000:.2f}ms'
    return response


def filter_fields(data, allowed_fields):
    """
    Filter dictionary to only include allowed fields (reduces JSON payload size)
    Works with nested dicts and lists of dicts
    """
    if isinstance(data, dict):
        return {k: filter_fields(v, allowed_fields) for k, v in data.items() if k in allowed_fields}
    elif isinstance(data, list):
        return [filter_fields(item, allowed_fields) for item in data]
    return data


def compact_weather_data(cities_weather):
    """
    Reduce weather data payload by removing redundant fields and rounding numbers
    Reduces JSON size by ~40%
    """
    compact = []
    for city in cities_weather:
        compact.append({
            'id': city.get('city_id'),
            'n': city.get('city_name'),
            's': city.get('state'),
            'lat': city.get('lat'),  # Needed for map
            'lon': city.get('lon'),  # Needed for map
            't': round(city.get('temperature', 0), 1) if city.get('temperature') else None,
            'dt': round(city.get('day_temp', 0), 1) if city.get('day_temp') else None,
            'nt': round(city.get('night_temp', 0), 1) if city.get('night_temp') else None,
            'h': round(city.get('humidity', 0), 1) if city.get('humidity') else None,
            'di': round(city.get('demand_index', 0), 1) if city.get('demand_index') else None,
            'dz': city.get('demand_zone'),
            'icon': city.get('zone_icon'),
            'ts': city.get('timestamp', '')[-8:] if city.get('timestamp') else ''  # Only send HH:MM:SS part
        })
    return compact


@app.after_request
def compress_response(response):
    """Gzip compress JSON/HTML/text responses > 300 bytes for faster transfers"""
    if (response.status_code < 200 or response.status_code >= 300 or
        response.direct_passthrough or
        'Content-Encoding' in response.headers or
        'gzip' not in request.headers.get('Accept-Encoding', '')):
        return response

    content_type = response.content_type or ''
    if not any(ct in content_type for ct in ['application/json', 'text/html', 'text/css', 'application/javascript', 'text/plain', 'application/xml']):
        return response

    data = response.get_data()
    if len(data) < 300:
        return response

    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode='wb', compresslevel=9) as gz:
        gz.write(data)
    compressed = buf.getvalue()

    # Only use compression if it actually reduces size
    if len(compressed) < len(data):
        response.set_data(compressed)
        response.headers['Content-Encoding'] = 'gzip'
        response.headers['Content-Length'] = len(compressed)
        response.headers['Vary'] = 'Accept-Encoding'
    return response


@app.after_request
def add_cache_headers(response):
    """Add browser cache headers to reduce redundant API calls"""
    path = request.path
    if response.status_code != 200:
        return response

    # Static files: cache aggressively (1 year with versioning)
    if path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
        response.headers['Expires'] = (datetime.now() + timedelta(days=365)).strftime('%a, %d %b %Y %H:%M:%S GMT')
        return response

    # Login page: never cache — prevents overlay state from persisting across page loads
    if path == '/login':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        return response

    # Main dashboard: never cache — must always hit the server for login_required check
    if path == '/':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        return response

    # Other HTML pages: short cache (5 min) for quick updates
    if path.endswith('.html'):
        response.headers['Cache-Control'] = 'private, max-age=300'
        return response

    # Weather API endpoints: cache 3 hours — data pulled once every 3 hours per client requirement
    # Note: dashboard-init is NOT cached so the browser immediately detects when the backend is offline
    if path in ('/api/weather/current', '/api/kpis'):
        response.headers['Cache-Control'] = 'private, max-age=10800'
        return response

    # dashboard-init must never be served from browser cache — always hit the server
    if path == '/api/dashboard-init':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        return response

    # Slow-changing data: cache 3 hours
    if path in ('/api/dsb-overview', '/api/demand-prediction', '/api/energy-estimates',
                '/api/service-predictions', '/api/weekly-summary', '/api/insights/simple'):
        response.headers['Cache-Control'] = 'private, max-age=10800'
        return response

    # Historical/monthly/heatmap/forecast data: cache 3 hours
    if any(path.startswith(p) for p in ['/api/historical', '/api/heatmap', '/api/monthly', '/api/forecast']):
        response.headers['Cache-Control'] = 'private, max-age=10800'
        return response

    # Default: no cache for other API endpoints
    return response


def login_required(f):
    """Decorator to require Supabase JWT login for a route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Allow preflight requests
        if request.method == 'OPTIONS':
            return f(*args, **kwargs)

        # Check for Authorization header first (for API calls)
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            try:
                # Basic validation using Supabase client
                if supabase_handler.enabled:
                    supabase_handler.client.auth.get_user(token)
                    return f(*args, **kwargs)
            except Exception as e:
                print(f"[Auth API] Invalid token: {e}")
                return jsonify({'error': 'Unauthorized', 'message': 'Invalid token'}), 401

        # Fallback to session check for traditional web routes
        # Also validate boot token so sessions from previous app runs are rejected
        if not session.get('logged_in') or session.get('boot_token') != BOOT_TOKEN:
            session.clear()
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

# Initialize services
weather_service = WeatherService()
alert_engine = AlertEngine()
data_processor = DataProcessor()
supabase_handler = SupabaseHandler()

# Initialize file-based persistent cache (survives restarts)
# TTL = 3 hours — data is pulled at most once every 3 hours per client requirement
file_cache = FileCache(cache_dir='cache_data', ttl_hours=3)
print("[FileCache] Persistent cache initialized (TTL=3 hours)")

# Simple in-memory cache for faster responses
cache = {
    'weather_data': None,
    'weather_timestamp': 0,
    'weather_data_stale': False,
    'weather_data_age': 0,
    'alerts_data': None,
    'alerts_timestamp': 0,
    'monthly_data': {},  # Cache monthly data by city_id and year
    'monthly_timestamp': {},
    'forecast_data': {},  # Cache forecast data by city_id
    'forecast_timestamp': {}
}
CACHE_TTL = 10800  # 3 hours — data pulled at most once every 3 hours per client requirement
MONTHLY_CACHE_TTL = 10800  # 3 hours cache for monthly data
FORECAST_CACHE_TTL = 10800  # 3 hours cache for seasonal forecast
ALERTS_TTL = 10800  # 3 hours — refresh alerts in sync with weather cadence

# Cache monitoring counters
cache_stats = {
    'cache_hits': 0,
    'cache_misses': 0,
    'api_calls': 0,
    'api_errors': 0,
    'total_api_latency': 0.0,
    'api_call_count': 0,
    'last_refresh': None,
    'sse_clients': 0
}

# SSE (Server-Sent Events) client management
sse_clients = []
sse_clients_lock = threading.Lock()

# Lock to prevent concurrent weather fetches (scheduler + user request race)
_weather_fetch_lock = threading.Lock()


def notify_sse_clients(event_type):
    """Push an event to all connected SSE clients."""
    message = f"event: {event_type}\ndata: {json_module.dumps({'timestamp': datetime.now().isoformat()})}\n\n"
    with sse_clients_lock:
        dead = []
        for q in sse_clients:
            try:
                q.put_nowait(message)
            except queue.Full:
                dead.append(q)
        for q in dead:
            sse_clients.remove(q)
        cache_stats['sse_clients'] = len(sse_clients)


def _safe_temp(city, key, fallback=None):
    """Get a temperature value that might be None. Returns fallback if None."""
    val = city.get(key)
    if val is not None:
        return val
    # Try deriving from base temperature
    temp = city.get('temperature')
    if temp is None:
        return fallback
    if key == 'night_temp':
        return temp - 5
    if key == 'day_temp':
        return temp
    return fallback


# Acknowledged alerts store
if 'alerts_ack' not in cache:
    cache['alerts_ack'] = set()


def get_cached_monthly_data(city_id, year):
    """
    Get monthly data with multi-tier caching:
    1. In-memory cache (1 hour TTL)
    2. File cache (12 hour TTL) - persists across restarts
    3. API fetch (only if both caches are stale/missing)
    """
    cache_key = f"{city_id}_{year}"
    now = time.time()

    # Tier 1: Check in-memory cache
    if (cache_key in cache['monthly_data'] and
        cache_key in cache['monthly_timestamp'] and
        (now - cache['monthly_timestamp'].get(cache_key, 0)) < MONTHLY_CACHE_TTL):
        cache_stats['cache_hits'] += 1
        return cache['monthly_data'][cache_key]

    # Tier 2: Check file cache
    file_cached_data = file_cache.get(f'monthly_{cache_key}')
    if file_cached_data:
        cache_stats['cache_hits'] += 1
        print(f"[Cache] Using file cache for monthly data: {city_id} ({year})")
        cache['monthly_data'][cache_key] = file_cached_data
        cache['monthly_timestamp'][cache_key] = now
        return file_cached_data

    # Tier 3: Fetch from API
    cache_stats['cache_misses'] += 1
    try:
        t0 = time.time()
        data = weather_service.get_monthly_averages(city_id, year)
        cache_stats['api_calls'] += 1
        cache_stats['total_api_latency'] += time.time() - t0
        cache_stats['api_call_count'] += 1
        cache['monthly_data'][cache_key] = data
        cache['monthly_timestamp'][cache_key] = now
        file_cache.set(f'monthly_{cache_key}', data)
        return data
    except Exception as e:
        cache_stats['api_errors'] += 1
        print(f"[Cache] Monthly data fetch failed for {city_id}/{year}: {e}")
        if cache_key in cache['monthly_data']:
            return cache['monthly_data'][cache_key]
        return []


def get_cached_forecast_fast(city_id, days=120):
    """Get forecast from cache only (no API call). Returns [] if not cached."""
    cache_key = f"{city_id}_{days}"
    if cache_key in cache['forecast_data']:
        return cache['forecast_data'][cache_key]
    data = file_cache.get(f'forecast_{cache_key}')
    if data:
        cache['forecast_data'][cache_key] = data
        cache['forecast_timestamp'][cache_key] = time.time()
        return data
    return []


def get_cached_forecast(city_id, days=120):
    """
    Get forecast data with multi-tier caching:
    1. In-memory cache (30 min TTL)
    2. File cache (12 hour TTL) - persists across restarts
    3. API fetch (only if both caches are stale/missing)
    """
    cache_key = f"{city_id}_{days}"
    now = time.time()

    # Tier 1: Check in-memory cache
    if (cache_key in cache['forecast_data'] and
        cache_key in cache['forecast_timestamp'] and
        (now - cache['forecast_timestamp'].get(cache_key, 0)) < FORECAST_CACHE_TTL):
        cache_stats['cache_hits'] += 1
        return cache['forecast_data'][cache_key]

    # Tier 2: Check file cache
    file_cached_data = file_cache.get(f'forecast_{cache_key}')
    if file_cached_data:
        cache_stats['cache_hits'] += 1
        print(f"[Cache] Using file cache for forecast: {city_id} ({days} days)")
        cache['forecast_data'][cache_key] = file_cached_data
        cache['forecast_timestamp'][cache_key] = now
        return file_cached_data

    # Tier 3: Fetch from API
    cache_stats['cache_misses'] += 1
    try:
        t0 = time.time()
        data = weather_service.get_forecast(city_id, days)
        cache_stats['api_calls'] += 1
        cache_stats['total_api_latency'] += time.time() - t0
        cache_stats['api_call_count'] += 1
        cache['forecast_data'][cache_key] = data
        cache['forecast_timestamp'][cache_key] = now
        file_cache.set(f'forecast_{cache_key}', data)
        return data
    except Exception as e:
        cache_stats['api_errors'] += 1
        print(f"[Cache] Forecast fetch failed for {city_id}: {e}")
        if cache_key in cache['forecast_data']:
            return cache['forecast_data'][cache_key]
        return []


def _ensure_all_cities_present(cities_weather):
    """Ensure there is an entry for every city in Config.CITIES.

    NEW BEHAVIOR: do NOT use cached/supabase/monthly data to populate missing
    values. Only present real/live values; otherwise return null (None) for
    the weather fields so the frontend shows NaN/empty state.
    """
    now_dt = datetime.now()
    by_id = {c.get('city_id') or c.get('id'): c for c in (cities_weather or [])}
    result = []

    for cfg in Config.CITIES:
        cid = cfg['id']
        entry = by_id.get(cid)

        if entry:
            # Ensure required keys and derived metrics are present for live data
            temp = entry.get('temperature')
            day_temp = entry.get('day_temp') or temp
            night_temp = entry.get('night_temp') or (temp - 5 if temp is not None else None)
            humidity = entry.get('humidity', None)

            entry['day_temp'] = day_temp
            entry['night_temp'] = night_temp
            entry['demand_index'] = data_processor.calculate_demand_index(day_temp, night_temp, humidity) if (day_temp is not None or night_temp is not None) else None
            entry['ac_hours'] = data_processor.calculate_ac_hours(day_temp, night_temp) if (day_temp is not None or night_temp is not None) else None
            entry['dsb_zone'] = data_processor.get_dsb_zone(entry['demand_index']) if entry.get('demand_index') is not None else None
            entry['wet_bulb'] = data_processor.calculate_wet_bulb(day_temp, humidity) if (day_temp is not None or humidity is not None) else None
            entry['demand_zone'] = cfg.get('demand_zone', 'Unknown')
            entry['zone_icon'] = cfg.get('zone_icon', '📍')
            entry['zone_traits'] = cfg.get('zone_traits', '')
            result.append(entry)
        else:
            # No live data -> show explicit nulls (no cached/fallback values)
            empty = {
                'city_id': cid,
                'city_name': cfg['name'],
                'state': cfg.get('state'),
                'temperature': None,
                'day_temp': None,
                'night_temp': None,
                'humidity': None,
                'feels_like': None,
                'wind_speed': None,
                'timestamp': None,
                'source': None,
                'demand_index': None,
                'ac_hours': None,
                'dsb_zone': None,
                'wet_bulb': None,
                'demand_zone': cfg.get('demand_zone', 'Unknown'),
                'zone_icon': cfg.get('zone_icon', '📍'),
                'zone_traits': cfg.get('zone_traits', '')
            }
            result.append(empty)

    return result


def _process_weather_cities(cities_weather):
    """Enrich raw weather data with derived metrics (demand index, AC hours, etc.)."""
    for city in cities_weather:
        temp = city.get('temperature')
        day_temp = city.get('day_temp') or temp
        night_temp = city.get('night_temp') or (temp - 5 if temp is not None else None)
        city['day_temp'] = day_temp
        city['night_temp'] = night_temp
        humidity = city.get('humidity', 60)
        city['demand_index'] = data_processor.calculate_demand_index(day_temp, night_temp, humidity)
        city['ac_hours'] = data_processor.calculate_ac_hours(day_temp, night_temp)
        city['dsb_zone'] = data_processor.get_dsb_zone(city['demand_index'])
        city['wet_bulb'] = data_processor.calculate_wet_bulb(day_temp, humidity)
        city_config = next((c for c in Config.CITIES if c['id'] == city.get('city_id')), {})
        city['demand_zone'] = city_config.get('demand_zone', 'Unknown')
        city['zone_icon'] = city_config.get('zone_icon', '')
        city['zone_traits'] = city_config.get('zone_traits', '')


def get_cached_weather():
    """
    Get weather data with multi-tier caching:
    1. In-memory cache (10 min TTL)
    2. File cache (12 hour TTL) - persists across restarts
    3. API fetch (only if both caches are stale/missing)
    """
    now = time.time()

    # Tier 1: Check in-memory cache (fastest)
    if cache['weather_data'] and (now - cache['weather_timestamp']) < CACHE_TTL:
        cache_stats['cache_hits'] += 1
        cache['weather_data_stale'] = False
        cache['weather_data_age'] = now - cache['weather_timestamp']
        return cache['weather_data']

    # Tier 2: Check file cache (persists across restarts)
    file_cached_data = file_cache.get('weather_all_cities')
    if file_cached_data:
        cache_stats['cache_hits'] += 1
        print("[Cache] Weather loaded from file cache")
        cache['weather_data'] = file_cached_data
        cache['weather_timestamp'] = now
        cache['weather_data_stale'] = False
        cache['weather_data_age'] = 0
        return file_cached_data

    # Tier 3: No valid cache - check if we have stale in-memory data to return immediately
    stale_data = cache['weather_data']
    if stale_data:
        cache_stats['cache_misses'] += 1
        print("[Cache] Returning stale data, refreshing in background")
        cache['weather_data_stale'] = True
        cache['weather_data_age'] = now - cache['weather_timestamp']
        filled = _ensure_all_cities_present(stale_data)

        # Refresh in background thread
        def refresh_cache():
            try:
                t0 = time.time()
                fresh_data = weather_service.get_all_cities_current()
                cache_stats['api_calls'] += 1
                cache_stats['total_api_latency'] += time.time() - t0
                cache_stats['api_call_count'] += 1
                _process_weather_cities(fresh_data)
                processed_data = _ensure_all_cities_present(fresh_data)
                cache['weather_data'] = processed_data
                cache['weather_timestamp'] = time.time()
                cache['weather_data_stale'] = False
                cache['weather_data_age'] = 0
                file_cache.set('weather_all_cities', processed_data)
                cache_stats['last_refresh'] = datetime.now().isoformat()
                notify_sse_clients('weather_update')
            except Exception as e:
                cache_stats['api_errors'] += 1
                print(f"[Cache] Background refresh failed: {e}")

        threading.Thread(target=refresh_cache, daemon=True).start()
        return filled

    # No cache at all — fetch synchronously so the first request gets real data
    # Use lock to prevent scheduler + user request from fetching simultaneously
    if not _weather_fetch_lock.acquire(blocking=False):
        # Startup preload (or another request) is already fetching.
        # Instead of blocking the Flask worker, return empty immediately.
        # The SSE weather_update event will push fresh data when preload finishes.
        print("[Cache] Preload in progress — returning empty immediately (SSE will push update)")
        return _ensure_all_cities_present([])

    cache_stats['cache_misses'] += 1
    print("[Cache] No cache found, fetching synchronously (first load)...")

    try:
        t0 = time.time()
        cities_weather = weather_service.get_all_cities_current()
        cache_stats['api_calls'] += 1
        cache_stats['total_api_latency'] += time.time() - t0
        cache_stats['api_call_count'] += 1

        if not cities_weather:
            print("[Cache] Synchronous fetch returned no live city data.")
            return _ensure_all_cities_present([])

        _process_weather_cities(cities_weather)

        # Update both caches
        processed_data = _ensure_all_cities_present(cities_weather)
        cache['weather_data'] = processed_data
        cache['weather_timestamp'] = time.time()
        cache['weather_data_stale'] = False
        cache['weather_data_age'] = 0
        file_cache.set('weather_all_cities', processed_data)
        cache_stats['last_refresh'] = datetime.now().isoformat()

        if supabase_handler.enabled and Config.PERSIST_WEATHER_LOGS:
            try:
                supabase_handler.save_weather_logs_batch(cities_weather)
            except Exception as e:
                print(f"[Supabase] Weather batch persist error: {e}")

        print(f"[Cache] Initial fetch completed ({len(cities_weather)} cities)")
        return processed_data
    except Exception as e:
        cache_stats['api_errors'] += 1
        print(f"[Cache] Synchronous fetch failed: {e}")
        return _ensure_all_cities_present([])
    finally:
        _weather_fetch_lock.release()


def get_minimal_weather_data():
    """Return an empty list — synthetic fallback data disabled by design."""
    # IMPORTANT: synthetic/fallback weather data has been intentionally disabled.
    # This function returns an empty list so callers must handle missing live data
    # explicitly instead of relying on fabricated values.
    return []


def get_cached_alerts(cities_weather):
    """Get alerts data with caching"""
    now = time.time()
    if cache['alerts_data'] and (now - cache['alerts_timestamp']) < CACHE_TTL:
        return cache['alerts_data']
    
    alerts = alert_engine.get_all_alerts(cities_weather)
    cache['alerts_data'] = alerts
    cache['alerts_timestamp'] = now
    return alerts


def refresh_alerts(force=False):
    """Refresh alerts and store in cache."""
    now = time.time()
    if (not force and cache.get('alerts_timestamp', 0) and (now - cache['alerts_timestamp'] < ALERTS_TTL)):
        return cache.get('alerts_data', [])

    try:
        cities_weather = get_cached_weather()
        alerts = alert_engine.get_all_alerts(cities_weather)
        # Attach stable ids and ack flag
        ts = int(time.time())
        new_alerts_to_send = []
        for idx, a in enumerate(alerts):
            a_id = f"{a.get('city_id')}_{ts}_{idx}"
            a['id'] = a_id
            a['acknowledged'] = a_id in cache.get('alerts_ack', set())
            # if webhook configured and this alert is Amber/Red/Kerala special and not sent yet
            if (a.get('alert_level') in ('red', 'orange', 'kerala_special') and
                a_id not in cache.get('alerts_webhook_sent', set()) and
                getattr(Config, 'ALERT_WEBHOOK_URLS', [])):
                new_alerts_to_send.append(a)
        cache['alerts_data'] = alerts
        cache['alerts_timestamp'] = now

        # Persist critical alerts to Supabase (non-blocking)
        if supabase_handler.enabled:
            critical = [a for a in alerts if a.get('alert_level') in ('red', 'orange', 'kerala_special')]
            if critical:
                def _persist_alerts(alert_list):
                    for alert in alert_list:
                        try:
                            supabase_handler.save_alert(alert)
                        except Exception as e:
                            print(f"[Supabase] Alert persist error: {e}")
                threading.Thread(target=_persist_alerts, args=(critical,), daemon=True).start()

        # send webhooks for newly discovered alerts (non-blocking)
        if new_alerts_to_send:
            try:
                threading.Thread(target=send_alert_webhooks, args=(new_alerts_to_send,), daemon=True).start()
            except Exception as e:
                app.logger.error('Failed to start webhook thread: %s', e)

        return alerts
    except Exception as e:
        app.logger.error('Error refreshing alerts: %s', e)
        return []


def send_alert_webhooks(alerts_list):
    """Send alert notifications to configured webhook URLs"""
    if not getattr(Config, 'ALERT_WEBHOOK_URLS', []):
        return
    
    for alert in alerts_list:
        try:
            webhook_data = {
                'alert': alert,
                'timestamp': datetime.now().isoformat(),
                'city': alert.get('city_name'),
                'level': alert.get('alert_level'),
                'message': alert.get('recommendation', {}).get('action', 'Alert triggered')
            }
            
            for url in Config.ALERT_WEBHOOK_URLS:
                try:
                    response = requests.post(
                        url,
                        json=webhook_data,
                        headers=Config.ALERT_WEBHOOK_HEADERS,
                        timeout=Config.ALERT_WEBHOOK_TIMEOUT
                    )
                    if response.status_code == 200:
                        print(f"[Webhook] Alert sent for {alert.get('city_name')}")
                        # Mark as sent
                        if 'alerts_webhook_sent' not in cache:
                            cache['alerts_webhook_sent'] = set()
                        cache['alerts_webhook_sent'].add(alert['id'])
                    else:
                        print(f"[Webhook] Failed to send alert: {response.status_code}")
                except Exception as e:
                    print(f"[Webhook] Error sending alert: {e}")
        except Exception as e:
            print(f"[Webhook] Error processing alert: {e}")


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page — authenticated via Supabase"""
    if session.get('logged_in'):
        return redirect('/')

    if request.method == 'POST':
        email = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        # ── Dev-mode bypass ──────────────────────────────────────────────────
        # When FLASK_ENV=development, skip Supabase entirely and use the
        # credentials defined in .env (DEV_LOGIN_USER / DEV_LOGIN_PASSWORD).
        if app.config.get('FLASK_ENV') == 'development':
            dev_user = os.environ.get('DEV_LOGIN_USER', 'admin@test.com')
            dev_pass = os.environ.get('DEV_LOGIN_PASSWORD', 'admin123')
            if email == dev_user and password == dev_pass:
                session['logged_in'] = True
                session['user_email'] = dev_user
                session['boot_token'] = BOOT_TOKEN
                session.permanent = True
                print(f"[Auth] Dev login accepted for '{dev_user}'")
                return jsonify({'success': True, 'redirect': '/'})
            else:
                return jsonify({'success': False, 'message': 'Invalid email or password.'})
        # ────────────────────────────────────────────────────────────────────

        supabase_url = app.config.get('SUPABASE_URL', '')
        supabase_key = app.config.get('SUPABASE_KEY', '')

        if not supabase_url or not supabase_key:
            return jsonify({'success': False, 'message': 'Authentication service unavailable. Check Supabase configuration.'})

        try:
            auth_response = requests.post(
                f"{supabase_url}/auth/v1/token?grant_type=password",
                headers={
                    'apikey': supabase_key,
                    'Content-Type': 'application/json'
                },
                json={'email': email, 'password': password},
                timeout=10
            )
            data = auth_response.json()
            if auth_response.status_code == 200 and data.get('access_token'):
                session['logged_in'] = True
                session['user_email'] = data.get('user', {}).get('email', email)
                session['boot_token'] = BOOT_TOKEN
                session.permanent = True
                return jsonify({'success': True, 'redirect': '/'})
            else:
                msg = data.get('error_description') or data.get('msg') or 'Invalid email or password.'
                return jsonify({'success': False, 'message': msg})
        except Exception as e:
            print(f"[Auth] Login error: {e}")
            return jsonify({'success': False, 'message': 'Could not reach authentication service. Please try again.'})

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logout and redirect to login"""
    session.clear()
    return redirect('/login')


@app.route('/debug/config')
def debug_config():
    """Temporary: check if env vars are loaded (remove after confirming)"""
    return jsonify({
        'supabase_url_set': bool(app.config.get('SUPABASE_URL')),
        'supabase_key_set': bool(app.config.get('SUPABASE_KEY')),
        'secret_key_set': bool(app.config.get('SECRET_KEY')),
        'supabase_handler_enabled': supabase_handler.enabled,
        'flask_env': app.config.get('FLASK_ENV')
    })

@app.route('/')
@login_required
def index():
    """Main dashboard page"""
    # Generate cache version based on file modification time for cache busting
    import hashlib
    from pathlib import Path
    cache_version = 'dev'
    try:
        static_dir = Path(__file__).parent / 'static'
        # Hash dashboard.js (the actual file served) so changes always bust the cache
        js_file = static_dir / 'js' / 'dashboard.js'
        if js_file.exists():
            with open(js_file, 'rb') as f:
                cache_version = hashlib.md5(f.read()).hexdigest()[:8]
    except Exception:
        pass
    
    return render_template('index.html', cache_version=cache_version)


@app.route('/api/cities')
def get_cities():
    """Get list of all configured cities"""
    return jsonify({
        'status': 'success',
        'data': Config.CITIES
    })


@app.route('/api/dashboard-init')
def get_dashboard_init():
    """
    Combined endpoint: weather + alerts + KPIs in one call to reduce round-trips
    OPTIMIZED: Uses compact field names to reduce JSON payload size by ~40%
    """
    try:
        cities_weather = get_cached_weather()

        # If no live weather data is available (empty list OR all-None placeholders),
        # return a safe, well-formed empty payload
        has_live_data = cities_weather and any(
            c.get('temperature') is not None for c in cities_weather
        )
        if not has_live_data:
            # Server is warming up — tell the client so it can show a helpful message
            warming_up = _weather_fetch_lock.locked()
            return jsonify({
                'status': 'success',
                'data': {
                    'weather': [],
                    'alerts': [],
                    'kpis': {
                        'hottest_day_city': None,
                        'hottest_night_city': None,
                        'season_status': None,
                        'days_to_peak': None,
                        'city_temps': [],
                        'night_temp_range': None,
                        'day_temp_range': None
                    },
                    'timestamp': None,
                    'stale': warming_up,
                    'warming_up': warming_up,
                    'data_age_seconds': 0
                }
            })

        # Alerts (limit to 20 most recent to reduce payload)
        alerts = refresh_alerts()
        filtered_alerts = [a for a in alerts if not a.get('acknowledged', False)][:20]

        # KPIs (computed from cached data, no extra API calls)
        hottest_day_city = data_processor.find_hottest_city(cities_weather, by_night=False)
        hottest_night_city = data_processor.find_hottest_city(cities_weather, by_night=True)

        city_temps = []
        for c in cities_weather:
            day_t = c.get('day_temp') or c.get('temperature')
            night_t = c.get('night_temp') or (c['temperature'] - 5 if c.get('temperature') is not None else None)
            if day_t is None or night_t is None:
                continue
            city_temps.append({
                'n': c['city_name'],  # Compact field names
                'dt': round(day_t, 1),
                'nt': round(night_t, 1)
            })
        city_temps.sort(key=lambda x: x['nt'], reverse=True)

        # Use cache-only forecast lookup — never block dashboard-init with a slow API call
        chennai_forecast = get_cached_forecast_fast('chennai', days=120)
        days_to_peak = data_processor.calculate_days_to_peak(chennai_forecast)

        # Build KPIs safely — hottest city lookups may return None if all temps are null
        if hottest_day_city and hottest_night_city and city_temps:
            hottest_night_temp = hottest_night_city.get('night_temp') or (hottest_night_city.get('temperature', 0) - 5)
            hottest_day_temp = hottest_day_city.get('day_temp') or hottest_day_city.get('temperature', 0)
            season_status = data_processor.get_season_status(hottest_day_temp, hottest_night_temp)
            # Use compact field names (nt, dt) for city_temps
            night_temps = [c['nt'] for c in city_temps]
            day_temps = [c['dt'] for c in city_temps]
            kpis = {
                'hottest_day_city': {
                    'n': hottest_day_city['city_name'],
                    't': round(hottest_day_temp, 1),
                    'type': 'day'
                },
                'hottest_night_city': {
                    'n': hottest_night_city['city_name'],
                    't': round(hottest_night_temp, 1),
                    'type': 'night'
                },
                'season': season_status,
                'd2p': days_to_peak,
                'cities': city_temps,
                'night_range': {'min': round(min(night_temps), 1), 'max': round(max(night_temps), 1)},
                'day_range': {'min': round(min(day_temps), 1), 'max': round(max(day_temps), 1)}
            }
        else:
            kpis = {
                'hottest_day_city': None,
                'hottest_night_city': None,
                'season': None,
                'd2p': days_to_peak,
                'cities': [],
                'night_range': None,
                'day_range': None
            }

        # Use compact weather data format
        compact_weather = compact_weather_data(cities_weather)

        return jsonify({
            'status': 'success',
            'data': {
                'weather': compact_weather,
                'alerts': filtered_alerts,
                'kpis': kpis,
                'ts': cities_weather[0]['timestamp'] if cities_weather else None,
                'stale': cache.get('weather_data_stale', False),
                'age': round(cache.get('weather_data_age', 0))
            }
        })
    except KeyError as ke:
        import traceback
        print(f"[ERROR] KeyError in dashboard-init: {ke}")
        print(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'KeyError: {str(ke)}',
            'traceback': traceback.format_exc()
        }), 500
    except Exception as e:
        import traceback
        print(f"[ERROR] Exception in dashboard-init: {e}")
        print(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/weather/current')
def get_current_weather():
    """Get current weather for all cities (with caching)"""
    try:
        cities_weather = get_cached_weather()
        
        return jsonify({
            'status': 'success',
            'data': cities_weather,
            'timestamp': cities_weather[0]['timestamp'] if cities_weather else None,
            'cached': cache['weather_timestamp'] > 0,
            'stale': cache.get('weather_data_stale', False),
            'data_age_seconds': round(cache.get('weather_data_age', 0))
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/weather/city/<city_id>')
def get_city_weather(city_id):
    """Get detailed weather for a specific city"""
    try:
        current = weather_service.get_current_weather(city_id)
        forecast = weather_service.get_forecast(city_id, days=7)
        historical = weather_service.get_historical_data(city_id, days=30)
        
        if not current:
            return jsonify({
                'status': 'error',
                'message': 'City not found'
            }), 404
        
        # Calculate demand index
        current['demand_index'] = data_processor.calculate_demand_index(
            current['temperature'],
            current.get('humidity')
        )
        
        return jsonify({
            'status': 'success',
            'data': {
                'current': current,
                'forecast': forecast,
                'historical': historical
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/alerts', endpoint='alerts_get')
def get_alerts():
    """Get alerts for all cities (with optional include_ack param)"""
    try:
        include_ack = request.args.get('include_ack', 'false').lower() == 'true'
        # Refresh and get alerts
        alerts = refresh_alerts()
        filtered = alerts if include_ack else [a for a in alerts if not a.get('acknowledged', False)]
        return jsonify({
            'status': 'success',
            'data': {'alerts': filtered},
            'count': len(filtered)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/kpis')
def get_kpis():
    """
    Get Key Performance Indicators per guide:
    - Hottest City (Day)
    - Hottest Night City
    - Season Status
    - Days to Peak
    """
    try:
        # Use CACHED weather data instead of making fresh API calls
        cities_weather = get_cached_weather()
        
        # Find hottest cities
        hottest_day_city = data_processor.find_hottest_city(cities_weather, by_night=False)
        hottest_night_city = data_processor.find_hottest_city(cities_weather, by_night=True)
        
        # Per-city temperatures (no combined averages - they don't make sense across diverse cities)
        city_temps = []
        for c in cities_weather:
            day_t = c.get('day_temp') or c.get('temperature')
            night_t = c.get('night_temp') or (c['temperature'] - 5 if c.get('temperature') is not None else None)
            if day_t is None or night_t is None:
                continue
            city_temps.append({
                'name': c['city_name'],
                'day_temp': round(day_t, 1),
                'night_temp': round(night_t, 1)
            })
        city_temps.sort(key=lambda x: x['night_temp'], reverse=True)

        chennai_forecast = get_cached_forecast_fast('chennai', days=120)
        days_to_peak = data_processor.calculate_days_to_peak(chennai_forecast)

        if hottest_day_city and hottest_night_city and city_temps:
            hottest_night_temp = hottest_night_city.get('night_temp') or (hottest_night_city.get('temperature', 0) - 5)
            hottest_day_temp = hottest_day_city.get('day_temp') or hottest_day_city.get('temperature', 0)
            season_status = data_processor.get_season_status(hottest_day_temp, hottest_night_temp)
            night_temps = [c['night_temp'] for c in city_temps]
            day_temps = [c['day_temp'] for c in city_temps]
            kpis = {
                'hottest_day_city': {
                    'name': hottest_day_city['city_name'],
                    'temperature': hottest_day_temp,
                    'type': 'day'
                },
                'hottest_night_city': {
                    'name': hottest_night_city['city_name'],
                    'temperature': hottest_night_temp,
                    'type': 'night',
                    'priority_note': '⭐ Night temp drives demand!'
                },
                'season_status': season_status,
                'days_to_peak': days_to_peak,
                'city_temps': city_temps,
                'night_temp_range': {'min': round(min(night_temps), 1), 'max': round(max(night_temps), 1)},
                'day_temp_range': {'min': round(min(day_temps), 1), 'max': round(max(day_temps), 1)}
            }
        else:
            kpis = {
                'hottest_day_city': None, 'hottest_night_city': None,
                'season_status': None, 'days_to_peak': days_to_peak,
                'city_temps': [], 'night_temp_range': None, 'day_temp_range': None
            }
        
        return jsonify({
            'status': 'success',
            'data': kpis
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/wave-sequence')
def get_wave_sequence():
    """
    Get Market Wave Sequence Analysis:
    Wave 1 (NOW) → Wave 2 (+2 weeks) → Wave 3 (+6 weeks)
    Lead indicators → Building markets → Lag markets
    """
    try:
        # Get forecasts from cache only — never block wave-sequence with API calls
        all_forecasts = {}
        for c in Config.CITIES:
            all_forecasts[c['id']] = get_cached_forecast_fast(c['id'], days=16)
        
        # Analyze wave sequence
        wave_data = alert_engine.analyze_wave_sequence(all_forecasts)
        
        return jsonify({
            'status': 'success',
            'data': wave_data
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export dashboard data to Excel"""
    try:
        cities_weather = get_cached_weather()
        alerts = alert_engine.get_all_alerts(cities_weather)


        # Prepare export data (Wave data removed as it was unused and slow to fetch)
        export_data = data_processor.prepare_export_data(
            cities_weather,
            alerts,
            None # wave_data unused in export_to_excel
        )

        # Export to Excel
        filename = f"ForecastWell_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"exports/{filename}"

        # Create exports directory if it doesn't exist
        import os
        os.makedirs('exports', exist_ok=True)

        success = data_processor.export_to_excel(cities_weather, filepath)

        if success:
            return jsonify({
                'status': 'success',
                'filename': filename,
                'path': filepath
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Failed to export data'
            }), 500

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/download/<filename>')
def download_file(filename):
    """Download exported file"""
    try:
        import os
        from flask import send_file
        filepath = os.path.join('exports', filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True)
        else:
            return jsonify({'status': 'error', 'message': 'File not found'}), 404
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/export/alert-report', methods=['POST'])
def export_alert_report():
    """Export alert-specific report to Excel"""
    try:
        cities_weather = get_cached_weather()
        alerts = alert_engine.get_all_alerts(cities_weather)
        
        # Filter for critical alerts only
        critical_alerts = [a for a in alerts if a['alert_level'] in ['red', 'orange', 'kerala_special']]
        
        # Create Excel file
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Critical Alerts"
        
        # Header
        headers = ['City', 'Alert Level', 'Night Temp', 'Day Temp', 'Demand Index', 'DSB Zone', 'AC Hours', 'Recommendation']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for alert in critical_alerts:
            dsb = alert.get('dsb_zone', '')
            dsb_label = dsb.get('label', dsb.get('zone', str(dsb))) if isinstance(dsb, dict) else str(dsb)
            rec = alert.get('recommendation', {})
            rec_action = rec.get('action', str(rec)) if isinstance(rec, dict) else str(rec)
            ws.append([
                alert.get('city', ''),
                alert.get('alert_level', '').upper(),
                f"{alert.get('night_temp', '')}°C",
                f"{alert.get('day_temp', '')}°C",
                alert.get('demand_index', ''),
                dsb_label,
                alert.get('ac_hours_estimated', ''),
                rec_action
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f"Alert_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"exports/{filename}"
        
        import os
        os.makedirs('exports', exist_ok=True)
        wb.save(filepath)
        
        return jsonify({
            'status': 'success',
            'filename': filename,
            'alerts_count': len(critical_alerts)
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/export/forecast-report', methods=['POST'])
def export_forecast_report():
    """Export forecast report for all cities"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        
        # Create sheet for each city
        for city in Config.CITIES:  # All cities
            forecast = get_cached_forecast(city['id'], days=30)
            
            if not forecast:
                continue
            
            ws = wb.create_sheet(title=city['name'][:31])  # Excel sheet name limit
            
            # Header
            headers = ['Date', 'Day', 'Day Temp', 'Night Temp', 'Humidity', 'Wind Speed', 'Source']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for day in forecast:
                ws.append([
                    day['date'],
                    day['day'],
                    f"{day['day_temp']}°C",
                    f"{day['night_temp']}°C",
                    f"{day.get('humidity', 'N/A')}%",
                    f"{day.get('wind_speed', 'N/A')} km/h",
                    day.get('source', 'Unknown')
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Save file
        filename = f"Forecast_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"exports/{filename}"
        
        import os
        os.makedirs('exports', exist_ok=True)
        wb.save(filepath)
        
        return jsonify({
            'status': 'success',
            'filename': filename,
            'cities_count': len([s for s in wb.sheetnames])
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/alerts/city/<city_id>')
def get_city_alert(city_id):
    """Get alert details for a specific city"""
    try:
        current = weather_service.get_current_weather(city_id)
        
        if not current:
            return jsonify({
                'status': 'error',
                'message': 'City not found'
            }), 404
        
        alert = alert_engine.analyze_temperature(
            current.get('day_temp', current.get('temperature', 0)),
            current.get('night_temp', current.get('temperature', 0)),
            current['city_name'],
            city_id
        )
        
        return jsonify({
            'status': 'success',
            'data': alert
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/forecast/analysis/<city_id>')
def get_forecast_analysis(city_id):
    """Get forecast analysis and recommendations for a city"""
    try:
        forecast = weather_service.get_forecast(city_id, days=7)
        
        # Get city name
        city_config = next((c for c in Config.CITIES if c['id'] == city_id), None)
        if not city_config:
            return jsonify({
                'status': 'error',
                'message': 'City not found'
            }), 404
        
        analysis = alert_engine.analyze_forecast(forecast, city_config['name'])
        
        return jsonify({
            'status': 'success',
            'data': analysis
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/dashboard/summary')
def get_dashboard_summary():
    """Get overall dashboard summary"""
    try:
        cities_weather = get_cached_weather()
        alerts = get_cached_alerts(cities_weather)

        # Calculate summary statistics
        temperatures = [c['temperature'] for c in cities_weather if c.get('temperature') is not None]
        if not temperatures:
            return jsonify({'status': 'success', 'data': {'total_cities': 0}})
        demand_indices = [data_processor.calculate_demand_index(c.get('day_temp') or c['temperature'], c.get('night_temp'), c.get('humidity'))
                         for c in cities_weather if c.get('temperature') is not None]
        
        critical_alerts = [a for a in alerts if a['alert_level'] == 'critical']
        high_alerts = [a for a in alerts if a['alert_level'] == 'high']
        
        summary = {
            'total_cities': len(cities_weather),
            'avg_temperature': round(sum(temperatures) / len(temperatures), 1),
            'max_temperature': round(max(temperatures), 1),
            'min_temperature': round(min(temperatures), 1),
            'avg_demand_index': round(sum(demand_indices) / len(demand_indices), 1),
            'critical_alerts': len(critical_alerts),
            'high_alerts': len(high_alerts),
            'overall_status': 'critical' if critical_alerts else ('high' if high_alerts else 'normal')
        }
        
        return jsonify({
            'status': 'success',
            'data': summary
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """Import weather data from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file uploaded'
            }), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({
                'status': 'error',
                'message': 'No file selected'
            }), 400

        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            import os
            from werkzeug.utils import secure_filename

            # Create uploads directory if it doesn't exist
            upload_dir = 'uploads'
            os.makedirs(upload_dir, exist_ok=True)

            # Save the file temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(upload_dir, filename)
            file.save(filepath)

            # Process the Excel file using the data processor
            success, message = data_processor.load_and_process_excel_data(filepath)

            if success:
                # Clean up the temporary file
                os.remove(filepath)

                return jsonify({
                    'status': 'success',
                    'message': message
                })
            else:
                return jsonify({
                    'status': 'error',
                    'message': f"Failed to process Excel file: {message}"
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': 'Invalid file format. Please upload .xlsx or .xls file'
            }), 400

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/health')
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'ForecastWell Dashboard',
        'version': '1.0.0'
    })


@app.route('/api/insights/simple')
def get_simple_insights():
    """Get basic AI-powered insights and recommendations"""
    try:
        cities_weather = get_cached_weather()
        alerts = get_cached_alerts(cities_weather)
        
        # Per-city metrics (no combined averages across diverse cities)
        hot_night_cities = []
        for c in cities_weather:
            day_t = c.get('day_temp') or c.get('temperature')
            night_t = c.get('night_temp') or (c['temperature'] - 5 if c.get('temperature') is not None else None)
            if day_t is None or night_t is None:
                continue
            hot_night_cities.append({
                'name': c['city_name'],
                'night_temp': round(night_t, 1),
                'day_temp': round(day_t, 1)
            })
        hot_night_cities.sort(key=lambda x: x['night_temp'], reverse=True)
        valid_cities = [c for c in cities_weather if c.get('humidity') is not None]
        avg_humidity = sum(c.get('humidity', 60) for c in valid_cities) / len(valid_cities) if valid_cities else 60
        
        # Count cities by alert level
        critical_count = len([a for a in alerts if a['alert_level'] == 'critical'])
        high_count = len([a for a in alerts if a['alert_level'] == 'high'])
        
        # Count cities by night temp threshold
        cities_above_24 = [c for c in hot_night_cities if c['night_temp'] >= 24]
        cities_above_22 = [c for c in hot_night_cities if c['night_temp'] >= 22]
        
        # Generate insights
        insights = []
        
        # Night temperature insight - per city, not averaged
        if cities_above_24:
            city_list = ', '.join([f"{c['name']} ({c['night_temp']}°C)" for c in cities_above_24])
            insights.append({
                'type': 'critical',
                'icon': '🌙',
                'title': f'{len(cities_above_24)} Cities at Peak Night Temps',
                'description': f'{city_list} — expect maximum AC usage (12-16 hours/day)',
                'action': 'Maximize inventory allocation to these markets'
            })
        elif cities_above_22:
            city_list = ', '.join([f"{c['name']} ({c['night_temp']}°C)" for c in cities_above_22])
            insights.append({
                'type': 'high',
                'icon': '🌡️',
                'title': f'{len(cities_above_22)} Cities with High Night Temps',
                'description': f'{city_list} — strong demand expected',
                'action': 'Accelerate stock movement to these areas'
            })
        
        # Humidity impact insight
        if avg_humidity >= 70:
            insights.append({
                'type': 'warning',
                'icon': '💧',
                'title': 'High Humidity Alert',
                'description': f'Average humidity at {round(avg_humidity, 1)}% increases perceived heat and AC dependency',
                'action': 'Focus on dehumidifier-enabled AC models'
            })
        
        # Market opportunity insight
        if critical_count > 0:
            insights.append({
                'type': 'opportunity',
                'icon': '🎯',
                'title': f'{critical_count} Markets at Peak Demand',
                'description': 'Multiple cities showing extreme conditions',
                'action': 'Deploy emergency stock replenishment'
            })
        
        # Energy consumption estimate
        total_ac_hours = sum(c.get('ac_hours', 8) for c in cities_weather)
        avg_ac_hours = total_ac_hours / len(cities_weather)
        energy_estimate = round(avg_ac_hours * 1.5 * len(cities_weather), 1)  # kWh estimate
        
        insights.append({
            'type': 'info',
            'icon': '⚡',
            'title': 'Energy Consumption Estimate',
            'description': f'Estimated regional AC energy consumption: {energy_estimate} kWh/day',
            'action': f'Average AC runtime: {round(avg_ac_hours, 1)} hours per household'
        })
        
        return jsonify({
            'status': 'success',
            'data': {
                'insights': insights,
                'summary': {
                    'city_temps': hot_night_cities,
                    'hottest_night': hot_night_cities[0] if hot_night_cities else None,
                    'coolest_night': hot_night_cities[-1] if hot_night_cities else None,
                    'avg_humidity': round(avg_humidity, 1),
                    'total_cities': len(cities_weather),
                    'critical_markets': critical_count,
                    'high_demand_markets': high_count
                }
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/insights')
def get_business_insights_page():
    """
    Get comprehensive business insights based on historical and forecast data
    Returns actionable recommendations, market analysis, and demand predictions
    """
    try:
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year
        
        # Get cached weather data for all cities
        cities_weather = get_cached_weather()
        
        # Get 4-month forecast + historical data for all cities IN PARALLEL
        forecast_by_city = {}
        historical_by_city = {}
        
        def fetch_city_forecast(city_id):
            return city_id, get_cached_forecast(city_id, days=120)
        
        def fetch_city_historical(city_id, year):
            return city_id, year, get_cached_monthly_data(city_id, year)
        
        with ThreadPoolExecutor(max_workers=12) as executor:
            # Submit all forecast tasks
            forecast_futures = [executor.submit(fetch_city_forecast, c['id']) for c in Config.CITIES]
            # Submit all historical tasks
            hist_futures = [executor.submit(fetch_city_historical, c['id'], y) for c in Config.CITIES for y in [2024, 2025]]
            
            for f in forecast_futures:
                city_id, data = f.result()
                forecast_by_city[city_id] = data
            
            for f in hist_futures:
                city_id, year, data = f.result()
                if city_id not in historical_by_city:
                    historical_by_city[city_id] = {}
                historical_by_city[city_id][year] = data
        
        # Calculate key metrics
        # 1. Days to peak (from Chennai as reference)
        chennai_forecast = forecast_by_city.get('chennai', [])
        days_to_peak = data_processor.calculate_days_to_peak(chennai_forecast)
        
        # 2. YoY Temperature Change (current month comparison)
        yoy_temp_change = 0
        if current_month <= 12:
            prev_year_avg = []
            curr_year_avg = []
            for city_id, hist in historical_by_city.items():
                last_year_month = next((m for m in hist.get(2025, []) if m['month'] == current_month), None)
                if last_year_month and last_year_month.get('avg_night_temp'):
                    prev_year_avg.append(last_year_month['avg_night_temp'])
            
            for city in cities_weather:
                night_t = city.get('night_temp') if city.get('night_temp') is not None else (city['temperature'] - 5 if city.get('temperature') is not None else None)
                if night_t is not None:
                    curr_year_avg.append(night_t)
            
            if prev_year_avg and curr_year_avg:
                yoy_temp_change = round((sum(curr_year_avg) / len(curr_year_avg)) - (sum(prev_year_avg) / len(prev_year_avg)), 1)
        
        # 3. Calculate 4-month demand potential
        month_demands = []
        for month_offset in range(4):
            target_month = current_month + month_offset + 1
            if target_month > 12:
                break
            
            for city_id, forecast in forecast_by_city.items():
                month_data = [d for d in forecast if datetime.strptime(d['date'], '%Y-%m-%d').month == target_month]
                if month_data:
                    avg_night = sum(d['night_temp'] for d in month_data) / len(month_data)
                    demand = min(100, max(0, int((avg_night - 15) * 7)))
                    month_demands.append(demand)
        
        avg_demand_potential = round(sum(month_demands) / len(month_demands)) if month_demands else 50
        
        # 4. Find top priority market
        city_scores = []
        for city in cities_weather:
            forecast = forecast_by_city.get(city['city_id'], [])
            future_nights = [d['night_temp'] for d in forecast[:60]]  # Next 2 months
            avg_future_night = sum(future_nights) / len(future_nights) if future_nights else 20
            score = min(100, (avg_future_night - 15) * 6)
            city_scores.append({
                'city': city['city_name'],
                'score': score,
                'night_temp': round(avg_future_night, 1)
            })
        
        city_scores.sort(key=lambda x: x['score'], reverse=True)
        top_market = city_scores[0] if city_scores else {'city': 'Chennai', 'score': 70, 'night_temp': 25}
        
        # Generate strategic recommendations
        recommendations = generate_strategic_recommendations(
            cities_weather, forecast_by_city, historical_by_city, current_month
        )
        
        # Generate monthly forecast breakdown
        monthly_forecast = generate_monthly_forecast_breakdown(forecast_by_city, current_month)
        
        # Generate action checklist
        action_checklist = generate_action_checklist(
            days_to_peak, avg_demand_potential, city_scores, current_month
        )
        
        # Generate city-wise insights
        city_insights = generate_city_insights(cities_weather, forecast_by_city, historical_by_city)
        
        # Market matrix data
        market_matrix = []
        months = ['Feb', 'Mar', 'Apr', 'May', 'Jun']
        for city in Config.CITIES[:8]:  # Limit to 8 cities
            city_forecast = forecast_by_city.get(city['id'], [])
            monthly_demands = []
            for i, month_name in enumerate(months):
                target_month = current_month + i
                if target_month > 12:
                    break
                month_data = [d for d in city_forecast if datetime.strptime(d['date'], '%Y-%m-%d').month == target_month]
                if month_data:
                    avg_night = sum(d['night_temp'] for d in month_data) / len(month_data)
                    demand = min(100, max(0, int((avg_night - 15) * 7)))
                    monthly_demands.append(demand)
                else:
                    monthly_demands.append(50)
            
            market_matrix.append({
                'city': city['name'],
                'city_id': city['id'],
                'demands': monthly_demands,
                'avg_demand': round(sum(monthly_demands) / len(monthly_demands)) if monthly_demands else 50
            })
        
        return jsonify({
            'status': 'success',
            'data': {
                'hero_metrics': {
                    'days_to_peak': days_to_peak,
                    'peak_trend': 'heating' if days_to_peak < 60 else 'building',
                    'yoy_temp_change': yoy_temp_change,
                    'yoy_trend': 'up' if yoy_temp_change > 0.5 else ('down' if yoy_temp_change < -0.5 else 'stable'),
                    'demand_potential': avg_demand_potential,
                    'top_market': top_market['city'],
                    'top_market_reason': f"{top_market['night_temp']}°C nights"
                },
                'recommendations': recommendations,
                'monthly_forecast': monthly_forecast,
                'action_checklist': action_checklist,
                'city_insights': city_insights,
                'market_matrix': {
                    'months': months[:len(monthly_forecast)],
                    'cities': market_matrix
                }
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/weekly-summary')
def get_weekly_summary():
    """Get weekly summary statistics using real forecast data"""
    try:
        cities_weather = get_cached_weather()
        
        weekly_data = []
        for city in cities_weather:
            city_id = city['city_id']
            # Get 7-day forecast (cache-only to avoid blocking)
            forecast = get_cached_forecast_fast(city_id, days=7)
            
            if forecast:
                avg_day = round(sum(d['day_temp'] for d in forecast) / len(forecast), 1)
                avg_night = round(sum(d['night_temp'] for d in forecast) / len(forecast), 1)
                max_temp = max(d['day_temp'] for d in forecast)
                min_night = min(d['night_temp'] for d in forecast)
                
                # Determine trend based on first vs last day
                start_temp = forecast[0]['day_temp']
                end_temp = forecast[-1]['day_temp']
                trend = 'rising' if end_temp > start_temp + 1 else ('falling' if end_temp < start_temp - 1 else 'stable')
            else:
                # Fallback to current if forecast fails (no random)
                avg_day = city.get('day_temp') or 35
                avg_night = city.get('night_temp') or 25
                max_temp = avg_day
                min_night = avg_night
                trend = 'stable'
            
            weekly_data.append({
                'city': city['city_name'],
                'city_id': city['city_id'],
                'avg_day_temp': avg_day,
                'avg_night_temp': avg_night,
                'max_temp': max_temp,
                'min_night_temp': min_night,
                'trend': trend,
                'demand_outlook': 'Very High' if avg_night >= 24 else ('High' if avg_night >= 22 else ('Moderate' if avg_night >= 20 else 'Low')),
                'demand_zone': city.get('demand_zone', 'Unknown')
            })
        
        return jsonify({
            'status': 'success',
            'data': weekly_data
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/forecast')
def get_forecast():
    """Get forecast data for a specific city. Query: city=<city_id>, days=<int>"""
    try:
        city_id = request.args.get('city')
        if not city_id:
            return jsonify({'status': 'error', 'message': 'City ID required'}), 400
            
        days = int(request.args.get('days', 30))
        forecast = get_cached_forecast(city_id, days=days)
        
        return jsonify({
            'status': 'success',
            'data': forecast,
            'city_id': city_id,
            'days': days
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/forecast/peak')
@login_required
def get_forecast_peak():
    """Find the peak temperature day in the full forecast window for a city.
    Always scans 120 days regardless of the currently selected forecast range.
    Query: city=<city_id> (required)
    Returns: peak date, peak day_temp, days_from_today, night_temp, humidity
    """
    city_id = request.args.get('city')
    if not city_id:
        return jsonify({'status': 'error', 'message': 'City ID required'}), 400

    try:
        forecast = get_cached_forecast(city_id, days=120)
        if not forecast:
            return jsonify({'status': 'error', 'message': 'No forecast data available'}), 404

        today = datetime.now().date()

        # Find the day with the highest day temperature
        peak_day = max(forecast, key=lambda d: d.get('day_temp') or d.get('temperature') or 0)

        peak_date_str = peak_day.get('date', '')
        peak_temp = peak_day.get('day_temp') or peak_day.get('temperature') or 0

        days_from_today = None
        if peak_date_str:
            try:
                peak_date = datetime.strptime(peak_date_str, '%Y-%m-%d').date()
                days_from_today = (peak_date - today).days
            except ValueError:
                pass

        return jsonify({
            'status': 'success',
            'city_id': city_id,
            'peak_date': peak_date_str,
            'peak_day_temp': round(peak_temp, 1),
            'peak_night_temp': round(peak_day.get('night_temp') or peak_temp - 5, 1),
            'peak_humidity': round(peak_day.get('humidity') or 50, 1),
            'days_from_today': days_from_today
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/forecast/year-compare')
@login_required
def get_forecast_year_compare():
    """Compare temperature for a selected date and city vs same date last year and 2 years ago.
    Query: city=<city_id> (required), date=YYYY-MM-DD (default: today)
    """
    city_id = request.args.get('city')
    if not city_id:
        return jsonify({'status': 'error', 'message': 'City ID required'}), 400

    city_config = next((c for c in Config.CITIES if c['id'] == city_id), None)
    if not city_config:
        return jsonify({'status': 'error', 'message': f'Unknown city: {city_id}'}), 404

    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        target_dt = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    today = datetime.now().date()
    last_year_str = target_dt.replace(year=target_dt.year - 1).strftime('%Y-%m-%d')
    two_years_ago_str = target_dt.replace(year=target_dt.year - 2).strftime('%Y-%m-%d')
    three_years_ago_str = target_dt.replace(year=target_dt.year - 3).strftime('%Y-%m-%d')

    def get_current_year_temp():
        """Get temperature for the target date via archive/current/forecast as appropriate"""
        try:
            if target_dt < today:
                return weather_service.get_daily_temp(city_id, date_str)
            if target_dt == today:
                current = weather_service.get_current_weather(city_id)
                if current:
                    day_t = current.get('day_temp') or current.get('temperature') or 0
                    night_t = current.get('night_temp') or (current.get('temperature', 0) - 5)
                    return {
                        'avg_temp': round((day_t + night_t) / 2, 1),
                        'day_temp': round(day_t, 1),
                        'night_temp': round(night_t, 1),
                    }
            # Future date — use forecast
            days_ahead = (target_dt - today).days
            if days_ahead <= 16:
                forecast = weather_service.get_forecast(city_id, days=days_ahead + 2)
                for day in forecast:
                    if day.get('date') == date_str:
                        day_t = day.get('day_temp') or 0
                        night_t = day.get('night_temp') or 0
                        return {
                            'avg_temp': round((day_t + night_t) / 2, 1),
                            'day_temp': round(day_t, 1),
                            'night_temp': round(night_t, 1),
                        }
        except Exception:
            pass
        return None

    def compare_temps(curr, hist):
        if not curr or not hist:
            return 'unknown'
        c, h = curr.get('avg_temp'), hist.get('avg_temp')
        if c is None or h is None:
            return 'unknown'
        diff = c - h
        return 'warmer' if diff > 0.5 else ('cooler' if diff < -0.5 else 'similar')

    try:
        # Fetch all four dates in parallel (3-year historical comparison)
        with ThreadPoolExecutor(max_workers=4) as ex:
            f_curr = ex.submit(get_current_year_temp)
            f_ly = ex.submit(weather_service.get_daily_temp, city_id, last_year_str)
            f_tya = ex.submit(weather_service.get_daily_temp, city_id, two_years_ago_str)
            f_3ya = ex.submit(weather_service.get_daily_temp, city_id, three_years_ago_str)
            current_data = f_curr.result()
            last_year_data = f_ly.result()
            two_years_ago_data = f_tya.result()
            three_years_ago_data = f_3ya.result()

        def build_hist(hist_data, hist_date_str):
            if hist_data:
                return {'date': hist_date_str, **hist_data,
                        'comparison': compare_temps(current_data, hist_data)}
            return {'date': hist_date_str, 'comparison': 'unavailable'}

        return jsonify({
            'status': 'success',
            'date': date_str,
            'last_year_date': last_year_str,
            'two_years_ago_date': two_years_ago_str,
            'three_years_ago_date': three_years_ago_str,
            'city': {
                'city_id': city_id,
                'city_name': city_config['name'],
                'current': current_data,
                'last_year': build_hist(last_year_data, last_year_str),
                'two_years_ago': build_hist(two_years_ago_data, two_years_ago_str),
                'three_years_ago': build_hist(three_years_ago_data, three_years_ago_str),
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/history')
def get_history():
    """Get daily historical data. Query: city=<city_id>, days=<int>"""
    try:
        city_id = request.args.get('city')
        days = int(request.args.get('days', 30))
        if not city_id:
             return jsonify({'status': 'error', 'message': 'City ID required'}), 400
             
        data = weather_service.get_historical_data(city_id, days=days)
        
        return jsonify({
            'status': 'success',
            'data': data
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/monthly-history')
def get_monthly_history():
    """Get monthly historical data for a specific year. Query: city=<city_id>, year=<int>"""
    try:
        city_id = request.args.get('city', 'chennai')
        year = int(request.args.get('year', 2024))
        
        data = get_cached_monthly_data(city_id, year)
        
        return jsonify({
            'status': 'success',
            'data': data,
            'city_id': city_id,
            'year': year
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/demand-prediction')
def get_demand_prediction():
    """Get demand prediction with confidence levels"""
    try:
        cities_weather = get_cached_weather()
        
        predictions = []
        for city in cities_weather:
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            humidity = city.get('humidity') or 60

            # Calculate demand score (0-100)
            # Night temp contributes 60%, Day temp 25%, Humidity 15%
            night_score = min(100, max(0, (night_temp - 15) * 10))
            day_score = min(100, max(0, (day_temp - 25) * 7.5))
            humidity_score = min(100, max(0, (humidity - 40) * 1.5))
            
            demand_score = round(night_score * 0.6 + day_score * 0.25 + humidity_score * 0.15, 1)
            
            # Calculate confidence based on data consistency
            confidence = 'High' if city.get('source', '') != 'Simulated' else 'Medium'
            
            # Determine demand level
            if demand_score >= 80:
                demand_level = 'Critical'
                recommendation = 'Maximum inventory allocation'
            elif demand_score >= 60:
                demand_level = 'High'
                recommendation = 'Increase stock levels'
            elif demand_score >= 40:
                demand_level = 'Moderate'
                recommendation = 'Maintain current levels'
            else:
                demand_level = 'Low'
                recommendation = 'Reduce inventory focus'
            
            predictions.append({
                'city': city['city_name'],
                'city_id': city['city_id'],
                'demand_score': demand_score,
                'demand_level': demand_level,
                'confidence': confidence,
                'recommendation': recommendation,
                'factors': {
                    'night_temp_contribution': round(night_score * 0.6, 1),
                    'day_temp_contribution': round(day_score * 0.25, 1),
                    'humidity_contribution': round(humidity_score * 0.15, 1)
                }
            })
        
        # Sort by demand score
        predictions.sort(key=lambda x: x['demand_score'], reverse=True)
        
        return jsonify({
            'status': 'success',
            'data': predictions
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/energy-estimates')
def get_energy_estimates():
    """Get energy consumption estimates by city"""
    try:
        cities_weather = get_cached_weather()
        
        estimates = []
        for city in cities_weather:
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            
            # Calculate AC hours based on temperature
            day_ac_hours = max(0, min(12, (day_temp - 28) * 1.2))
            night_ac_hours = max(0, min(12, (night_temp - 20) * 3))
            total_ac_hours = round(day_ac_hours + night_ac_hours, 1)
            
            # Energy consumption (assuming 1.5 kW average AC)
            daily_kwh = round(total_ac_hours * 1.5, 1)
            monthly_kwh = round(daily_kwh * 30, 1)
            monthly_cost = round(monthly_kwh * 6.5, 0)  # Avg ₹6.5/kWh
            
            estimates.append({
                'city': city['city_name'],
                'city_id': city['city_id'],
                'ac_hours_day': round(day_ac_hours, 1),
                'ac_hours_night': round(night_ac_hours, 1),
                'total_ac_hours': total_ac_hours,
                'daily_kwh': daily_kwh,
                'monthly_kwh': monthly_kwh,
                'estimated_monthly_cost': f'₹{monthly_cost:,.0f}'
            })
        
        return jsonify({
            'status': 'success',
            'data': estimates
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/comparison/historical')
def get_historical_comparison():
    """
    Get historical comparison data.
    Previously simulated; now returns empty state until real historical aggregate data is available.
    """
    return jsonify({
        'status': 'success',
        'data': []
    })


@app.route('/api/historical/date-compare')
def get_date_comparison():
    """
    Compare weather on a specific date across current year, last year, and 2 years ago.
    Uses Open-Meteo Archive API for real historical data.
    Query params: date (YYYY-MM-DD), city (city_id, default: all)
    """
    try:
        date_str = request.args.get('date')
        city_filter = request.args.get('city', 'all')

        if not date_str:
            return jsonify({'status': 'error', 'message': 'date parameter required (YYYY-MM-DD)'}), 400

        target_date = datetime.strptime(date_str, '%Y-%m-%d')
        today = datetime.now()

        # Build the 3 comparison dates: same month-day for target year, year-1, year-2
        target_year = target_date.year
        years = [target_year, target_year - 1, target_year - 2]

        # Don't request future dates from archive API
        comparison_dates = []
        for y in years:
            try:
                d = datetime(y, target_date.month, target_date.day)
            except ValueError:
                # handle Feb 29 in non-leap years
                d = datetime(y, target_date.month, 28)
            # Only include if date is in the past (archive API only has past data)
            if d.date() < today.date():
                comparison_dates.append(d)
            else:
                comparison_dates.append(None)

        # Determine which cities to query
        if city_filter and city_filter != 'all':
            cities_to_query = [c for c in Config.CITIES if c['id'] == city_filter]
        else:
            cities_to_query = Config.CITIES

        results = []
        for city in cities_to_query:
            city_result = {
                'city_id': city['id'],
                'city_name': city['name'],
                'state': city['state'],
                'years': []
            }

            for i, comp_date in enumerate(comparison_dates):
                year_label = years[i]
                if comp_date is None:
                    # Future date - use current weather or forecast
                    current = weather_service.get_current_weather(city['id'])
                    if current:
                        city_result['years'].append({
                            'year': year_label,
                            'date': target_date.strftime('%Y-%m-%d'),
                            'day_temp': _safe_temp(current, 'day_temp', 32),
                            'night_temp': _safe_temp(current, 'night_temp', 20),
                            'humidity': current.get('humidity') or 65,
                            'source': 'Current/Forecast',
                            'is_current': True
                        })
                    continue

                # Fetch from Open-Meteo Archive API
                try:
                    params = {
                        'latitude': city['lat'],
                        'longitude': city['lon'],
                        'start_date': comp_date.strftime('%Y-%m-%d'),
                        'end_date': comp_date.strftime('%Y-%m-%d'),
                        'daily': 'temperature_2m_max,temperature_2m_min,relative_humidity_2m_mean',
                        'timezone': 'Asia/Kolkata'
                    }
                    resp = requests.get(
                        'https://archive-api.open-meteo.com/v1/archive',
                        params=params, timeout=10
                    )
                    if resp.status_code == 200:
                        data = resp.json()
                        daily = data.get('daily', {})
                        if daily.get('temperature_2m_max') and daily['temperature_2m_max'][0] is not None:
                            city_result['years'].append({
                                'year': year_label,
                                'date': comp_date.strftime('%Y-%m-%d'),
                                'day_temp': round(daily['temperature_2m_max'][0], 1),
                                'night_temp': round(daily['temperature_2m_min'][0], 1),
                                'humidity': round(daily['relative_humidity_2m_mean'][0], 1) if daily.get('relative_humidity_2m_mean') and daily['relative_humidity_2m_mean'][0] else None,
                                'source': 'Open-Meteo Archive',
                                'is_current': False
                            })
                        else:
                            city_result['years'].append({
                                'year': year_label,
                                'date': comp_date.strftime('%Y-%m-%d'),
                                'day_temp': None,
                                'night_temp': None,
                                'humidity': None,
                                'source': 'No data available',
                                'is_current': False
                            })
                    else:
                        city_result['years'].append({
                            'year': year_label,
                            'date': comp_date.strftime('%Y-%m-%d'),
                            'day_temp': None, 'night_temp': None, 'humidity': None,
                            'source': f'API Error {resp.status_code}',
                            'is_current': False
                        })
                except Exception as api_err:
                    city_result['years'].append({
                        'year': year_label,
                        'date': comp_date.strftime('%Y-%m-%d'),
                        'day_temp': None, 'night_temp': None, 'humidity': None,
                        'source': f'Error: {str(api_err)}',
                        'is_current': False
                    })

            # Calculate trend comparison
            valid_years = [y for y in city_result['years'] if y['day_temp'] is not None]
            if len(valid_years) >= 2:
                latest = valid_years[0]
                previous = valid_years[1]
                day_diff = round(latest['day_temp'] - previous['day_temp'], 1)
                night_diff = round(latest['night_temp'] - previous['night_temp'], 1)
                city_result['comparison'] = {
                    'day_diff': day_diff,
                    'night_diff': night_diff,
                    'trend': 'warmer' if (day_diff + night_diff) / 2 > 0 else ('cooler' if (day_diff + night_diff) / 2 < 0 else 'same'),
                    'day_label': f'+{day_diff}°C' if day_diff > 0 else f'{day_diff}°C',
                    'night_label': f'+{night_diff}°C' if night_diff > 0 else f'{night_diff}°C'
                }
            else:
                city_result['comparison'] = None

            results.append(city_result)

        return jsonify({
            'status': 'success',
            'data': results,
            'query': {
                'date': date_str,
                'city': city_filter,
                'years_compared': years
            }
        })
    except ValueError as ve:
        return jsonify({'status': 'error', 'message': f'Invalid date format: {str(ve)}'}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/historical/lookback-compare')
def get_lookback_compare():
    """
    For a selected date and city, return weather for:
      - the selected date
      - 7 days before
      - 14 days before
      - 1 month before
    Each entry is compared vs the same date 1 year ago (warmer / cooler / similar).
    Query params: date (YYYY-MM-DD), city (city_id)
    """
    try:
        date_str = request.args.get('date')
        city_id = request.args.get('city', 'chennai')

        if not date_str:
            return jsonify({'status': 'error', 'message': 'date parameter required (YYYY-MM-DD)'}), 400

        from datetime import date as date_type
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        today = datetime.now().date()

        # Build the 4 anchor dates
        one_month_ago = target_date.replace(
            year=target_date.year if target_date.month > 1 else target_date.year - 1,
            month=target_date.month - 1 if target_date.month > 1 else 12
        )
        anchor_dates = [
            ('Selected Date', target_date),
            ('7 Days Ago',    target_date - timedelta(days=7)),
            ('14 Days Ago',   target_date - timedelta(days=14)),
            ('1 Month Ago',   one_month_ago),
        ]

        cities_match = [c for c in Config.CITIES if c['id'] == city_id]
        if not cities_match:
            return jsonify({'status': 'error', 'message': f'City {city_id} not found'}), 404
        city = cities_match[0]

        def fetch_day(d):
            """Fetch daily weather for one date from Open-Meteo Archive API."""
            if d > today:
                return None
            try:
                params = {
                    'latitude': city['lat'],
                    'longitude': city['lon'],
                    'start_date': d.strftime('%Y-%m-%d'),
                    'end_date':   d.strftime('%Y-%m-%d'),
                    'daily': 'temperature_2m_max,temperature_2m_min,relative_humidity_2m_mean',
                    'timezone': 'Asia/Kolkata'
                }
                resp = requests.get(
                    'https://archive-api.open-meteo.com/v1/archive',
                    params=params, timeout=10
                )
                if resp.status_code == 200:
                    daily = resp.json().get('daily', {})
                    if daily.get('temperature_2m_max') and daily['temperature_2m_max'][0] is not None:
                        return {
                            'date': d.strftime('%Y-%m-%d'),
                            'day_temp': round(daily['temperature_2m_max'][0], 1),
                            'night_temp': round(daily['temperature_2m_min'][0], 1),
                            'humidity': round(daily['relative_humidity_2m_mean'][0], 1)
                                        if daily.get('relative_humidity_2m_mean') and daily['relative_humidity_2m_mean'][0]
                                        else None
                        }
            except Exception:
                pass
            return None

        # Build full list of (label, which, date) to fetch in parallel
        fetch_tasks = []
        for label, anchor in anchor_dates:
            try:
                prev_year_date = anchor.replace(year=anchor.year - 1)
            except ValueError:
                prev_year_date = anchor.replace(year=anchor.year - 1, day=28)
            fetch_tasks.append((label, 'current',   anchor))
            fetch_tasks.append((label, 'prev_year', prev_year_date))

        raw = {}
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(fetch_day, d): (lbl, which) for lbl, which, d in fetch_tasks}
            for future in futures:
                lbl, which = futures[future]
                if lbl not in raw:
                    raw[lbl] = {}
                raw[lbl][which] = future.result()

        results = []
        for label, anchor in anchor_dates:
            try:
                prev_year_date = anchor.replace(year=anchor.year - 1)
            except ValueError:
                prev_year_date = anchor.replace(year=anchor.year - 1, day=28)

            current  = raw.get(label, {}).get('current')
            prev_yr  = raw.get(label, {}).get('prev_year')

            entry = {
                'label':          label,
                'date':           anchor.strftime('%Y-%m-%d'),
                'prev_year_date': prev_year_date.strftime('%Y-%m-%d'),
                'current':        current,
                'prev_year':      prev_yr,
                'comparison':     None
            }

            if current and prev_yr:
                day_diff   = round(current['day_temp']   - prev_yr['day_temp'],   1)
                night_diff = round(current['night_temp'] - prev_yr['night_temp'], 1)
                avg_diff   = round((day_diff + night_diff) / 2, 1)
                trend = 'warmer' if avg_diff > 0.3 else ('cooler' if avg_diff < -0.3 else 'similar')
                entry['comparison'] = {
                    'day_diff':    day_diff,
                    'night_diff':  night_diff,
                    'avg_diff':    avg_diff,
                    'trend':       trend,
                    'day_label':   f'+{day_diff}°C' if day_diff > 0 else f'{day_diff}°C',
                    'night_label': f'+{night_diff}°C' if night_diff > 0 else f'{night_diff}°C',
                    'avg_label':   f'+{avg_diff}°C' if avg_diff > 0 else f'{avg_diff}°C',
                }

            results.append(entry)

        return jsonify({
            'status': 'success',
            'city':   {'id': city['id'], 'name': city['name'], 'state': city.get('state', '')},
            'data':   results
        })

    except ValueError as ve:
        return jsonify({'status': 'error', 'message': f'Invalid date: {str(ve)}'}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/historical/two-years')
def get_two_year_historical():
    """
    Get 2-year historical data from January 2024 to present (February 2026)
    Supports date range filtering via query parameters
    """
    try:
        # Get date range from query parameters
        start_date_str = request.args.get('start_date', '2023-01-01')
        end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
        city_filter = request.args.get('city', 'all')
        granularity = request.args.get('granularity', 'daily')  # daily, weekly, monthly

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            start_date = datetime(2023, 1, 1)
            end_date = datetime.now()

        # Ensure we're within Jan 2023 to present (3-year window)
        min_date = datetime(2023, 1, 1)
        if start_date < min_date:
            start_date = min_date
        if end_date > datetime.now():
            end_date = datetime.now()
        
        historical_data = generate_two_year_historical_data(
            start_date, end_date, city_filter, granularity
        )
        
        return jsonify({
            'status': 'success',
            'data': historical_data,
            'meta': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'city_filter': city_filter,
                'granularity': granularity,
                'total_records': len(historical_data['timeline']),
                'data_range': 'January 2023 - Present (3 Years)'
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/historical/summary')
def get_historical_summary():
    """
    Get summary statistics for the 2-year historical period.
    Currently mocked to return empty structure as real data analysis requires
    migrating full historical dataset to Supabase/DB.
    """
    return jsonify({
        'status': 'success',
        'data': {
            'period': {'start': 'N/A', 'end': 'N/A'},
            'temperature_stats': {},
            'demand_trends': {},
            'city_rankings': [],
            'seasonal_patterns': []
        }
    })


def generate_two_year_historical_data(start_date, end_date, city_filter, granularity):
    """
    Generate realistic 2-year historical weather data
    Based on actual South India climate patterns
    """
    # Seasonal temperature patterns for South India
    seasonal_patterns = {
        1: {'base_day': 30, 'base_night': 20, 'variation': 3},   # January
        2: {'base_day': 32, 'base_night': 21, 'variation': 3},   # February
        3: {'base_day': 35, 'base_night': 24, 'variation': 3},   # March
        4: {'base_day': 38, 'base_night': 27, 'variation': 4},   # April
        5: {'base_day': 40, 'base_night': 29, 'variation': 4},   # May
        6: {'base_day': 38, 'base_night': 27, 'variation': 3},   # June
        7: {'base_day': 34, 'base_night': 25, 'variation': 3},   # July (Monsoon)
        8: {'base_day': 33, 'base_night': 24, 'variation': 2},   # August
        9: {'base_day': 33, 'base_night': 24, 'variation': 2},   # September
        10: {'base_day': 32, 'base_night': 23, 'variation': 2},  # October
        11: {'base_day': 30, 'base_night': 21, 'variation': 2},  # November
        12: {'base_day': 29, 'base_night': 19, 'variation': 2},  # December
    }
    
    # City-specific temperature offsets
    city_offsets = {
        'chennai': {'day': 2, 'night': 3},      # Coastal, warmer nights
        'hyderabad': {'day': 1, 'night': -1},   # Moderate
        'bangalore': {'day': -3, 'night': -3},  # Cooler highland
        'visakhapatnam': {'day': 1, 'night': 2},# Coastal
        'coimbatore': {'day': -2, 'night': -2}, # Cooler
        'kochi': {'day': 0, 'night': 3},        # Coastal, warm nights
    }
    
    cities = Config.CITIES if city_filter == 'all' else [
        c for c in Config.CITIES if c['id'] == city_filter
    ]
    
    timeline = []
    city_data = {city['id']: [] for city in cities}
    
    current_date = start_date
    delta = timedelta(days=1)
    
    if granularity == 'weekly':
        delta = timedelta(days=7)
    elif granularity == 'monthly':
        delta = timedelta(days=30)
    
    while current_date <= end_date:
        month = current_date.month
        pattern = seasonal_patterns[month]
        
        # Add year-over-year warming trend (climate change effect, base 2023)
        year_offset = (current_date.year - 2023) * 0.5
        
        date_entry = {
            'date': current_date.strftime('%Y-%m-%d'),
            'month': current_date.strftime('%B %Y'),
            'week': current_date.strftime('Week %W, %Y'),
            'cities': {}
        }
        
        for city in cities:
            city_id = city['id']
            offset = city_offsets.get(city_id, {'day': 0, 'night': 0})
            
            # Calculate temperatures with realistic variations
            day_temp = pattern['base_day'] + offset['day'] + year_offset + random.uniform(-pattern['variation'], pattern['variation'])
            night_temp = pattern['base_night'] + offset['night'] + year_offset + random.uniform(-pattern['variation']/2, pattern['variation']/2)
            
            # Calculate demand index based on night temp (primary driver)
            demand_index = min(100, max(0, int((night_temp - 15) * 7)))
            
            # Calculate AC hours
            ac_hours = max(0, min(24, round((night_temp - 18) * 2.5 + (day_temp - 30) * 0.5, 1)))
            
            city_entry = {
                'day_temp': round(day_temp, 1),
                'night_temp': round(night_temp, 1),
                'humidity': round(50 + random.uniform(-10, 20), 1),
                'demand_index': demand_index,
                'ac_hours': ac_hours
            }
            
            date_entry['cities'][city_id] = city_entry
            city_data[city_id].append({
                'date': current_date.strftime('%Y-%m-%d'),
                **city_entry
            })
        
        timeline.append(date_entry)
        current_date += delta
    
    # Calculate per-city statistics (no combined averages across diverse cities)
    yearly_stats = {}
    for year in [2023, 2024, 2025, 2026]:
        year_data = [t for t in timeline if t['date'].startswith(str(year))]
        if year_data:
            # Per-city stats
            city_year_stats = {}
            for cid in [c['id'] for c in cities]:
                city_day_temps = []
                city_night_temps = []
                city_demands = []
                for entry in year_data:
                    if cid in entry['cities']:
                        cv = entry['cities'][cid]
                        city_day_temps.append(cv['day_temp'])
                        city_night_temps.append(cv['night_temp'])
                        city_demands.append(cv['demand_index'])
                if city_day_temps:
                    city_name = next((c['name'] for c in cities if c['id'] == cid), cid)
                    city_year_stats[cid] = {
                        'name': city_name,
                        'max_day_temp': round(max(city_day_temps), 1),
                        'max_night_temp': round(max(city_night_temps), 1),
                        'avg_demand': round(sum(city_demands) / len(city_demands), 1)
                    }
            
            # Find hottest/coolest cities for the year
            sorted_by_night = sorted(city_year_stats.values(), key=lambda x: x['max_night_temp'], reverse=True)
            
            yearly_stats[year] = {
                'city_stats': city_year_stats,
                'hottest_city': sorted_by_night[0] if sorted_by_night else None,
                'coolest_city': sorted_by_night[-1] if sorted_by_night else None,
                'peak_day_temp': round(max(cv['max_day_temp'] for cv in city_year_stats.values()), 1) if city_year_stats else 0,
                'peak_night_temp': round(max(cv['max_night_temp'] for cv in city_year_stats.values()), 1) if city_year_stats else 0,
                'data_points': len(year_data)
            }
    
    return {
        'timeline': timeline,
        'city_data': city_data,
        'yearly_stats': yearly_stats,
        'cities': [{'id': c['id'], 'name': c['name']} for c in cities]
    }


@app.route('/api/heatmap/monthly')
def get_heatmap_monthly_data():
    """
    Get monthly temperature data for heatmap from Open-Meteo API
    Returns real historical data for 2024, 2025 and current year
    """
    try:
        city_id = request.args.get('city', 'chennai')
        
        # Get city config
        city = next((c for c in Config.CITIES if c['id'] == city_id), None)
        if not city:
            city = Config.CITIES[0]  # Default to first city
            city_id = city['id']
        
        result = {
            'city_id': city_id,
            'city_name': city['name'],
            'years': {}
        }
        
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        # Fetch data for 2023, 2024, 2025, 2026 (3-year history + current)
        for year in [2023, 2024, 2025, 2026]:
            monthly_data = get_cached_monthly_data(city_id, year)
            
            result['years'][year] = []
            for month_data in monthly_data:
                avg_temp = None
                if month_data['avg_day_temp'] is not None:
                    # Use day temp as the heatmap value (peak temperature)
                    avg_temp = month_data['avg_day_temp']
                
                result['years'][year].append({
                    'month': month_data['month_name'],
                    'month_num': month_data['month'],
                    'temp': avg_temp,
                    'day_temp': month_data['avg_day_temp'],
                    'night_temp': month_data['avg_night_temp'],
                    'source': month_data['source'],
                    'is_forecast': month_data.get('is_forecast', False)
                })
        
        # Add forecast data for future months in current year
        # Open-Meteo Seasonal API provides up to 7 months forecast (using ECMWF SEAS5)
        # We request 4 months (120 days) of forecast data (cached for 30 mins)
        forecast_end_month = min(current_month + 4, 12)
        forecast = get_cached_forecast(city_id, days=120)  # 4 months forecast from Seasonal API
        
        if forecast:
            # Group forecast by month
            forecast_by_month = {}
            for day_data in forecast:
                date = datetime.strptime(day_data['date'], '%Y-%m-%d')
                month_num = date.month
                if month_num not in forecast_by_month:
                    forecast_by_month[month_num] = {'day_temps': [], 'night_temps': []}
                forecast_by_month[month_num]['day_temps'].append(day_data['day_temp'])
                forecast_by_month[month_num]['night_temps'].append(day_data['night_temp'])
            
            # Update 2026 data with forecast for future months
            for month_num in range(current_month + 1, forecast_end_month + 1):
                if month_num in forecast_by_month:
                    # Use actual forecast data from Seasonal API
                    temps = forecast_by_month[month_num]
                    avg_day = sum(temps['day_temps']) / len(temps['day_temps'])
                    avg_night = sum(temps['night_temps']) / len(temps['night_temps'])
                    
                    # Update the month in 2026 data
                    for month_data in result['years'][2026]:
                        if month_data['month_num'] == month_num:
                            month_data['temp'] = round(avg_day, 1)
                            month_data['day_temp'] = round(avg_day, 1)
                            month_data['night_temp'] = round(avg_night, 1)
                            month_data['source'] = 'Open-Meteo Seasonal (ECMWF SEAS5)'
                            month_data['is_forecast'] = True
                            break
                else:
                    # Use historical average from same month in 2025 as estimate
                    # (This provides a data-driven estimate for months beyond 16-day forecast)
                    historical_month = next((m for m in result['years'][2025] if m['month_num'] == month_num), None)
                    if historical_month and historical_month['temp']:
                        for month_data in result['years'][2026]:
                            if month_data['month_num'] == month_num:
                                # Add slight warming trend (+0.5°C year-over-year)
                                month_data['temp'] = round(historical_month['temp'] + 0.5, 1)
                                month_data['day_temp'] = round(historical_month['day_temp'] + 0.5, 1) if historical_month['day_temp'] else None
                                month_data['night_temp'] = round(historical_month['night_temp'] + 0.5, 1) if historical_month['night_temp'] else None
                                month_data['source'] = 'Historical Estimate (2025 + 0.5°C)'
                                month_data['is_forecast'] = True
                                break
        
        return jsonify({
            'status': 'success',
            'data': result,
            'meta': {
                'current_year': current_year,
                'current_month': current_month,
                'forecast_end_month': forecast_end_month,
                'note': 'Months 1-16 days ahead use Open-Meteo forecast; beyond that uses historical estimates'
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/comparison/monthly-yoy')
def get_monthly_yoy_comparison():
    """
    Get month-wise Year-over-Year comparison data from Open-Meteo API
    Compare same month across 2023, 2024, 2025, 2026 (3-year history + current)
    """
    try:
        city_filter = request.args.get('city', 'all')
        month_filter = request.args.get('month', None)  # 1-12, or None for all months

        cities = Config.CITIES
        if city_filter != 'all':
            cities = [c for c in cities if c['id'] == city_filter]

        # Use first city if filtering, otherwise use Chennai as representative
        sample_city = cities[0] if len(cities) == 1 else next((c for c in cities if c['id'] == 'chennai'), cities[0])

        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']

        comparison_data = []
        current_month = datetime.now().month
        current_year = datetime.now().year

        # Forecast extends +4 months from current month (e.g., Feb -> June)
        forecast_end_month = min(current_month + 4, 12)

        # Fetch monthly averages from Open-Meteo API for each year (cached for 1 hour)
        yearly_data = {}
        for year in [2023, 2024, 2025, 2026]:
            yearly_data[year] = get_cached_monthly_data(sample_city['id'], year)
        
        # Get forecast data for future months (up to 4 months, cached for 30 mins)
        forecast_data = get_cached_forecast(sample_city['id'], days=120)  # 4 months from Seasonal API
        forecast_by_month = {}
        if forecast_data:
            for day_data in forecast_data:
                date = datetime.strptime(day_data['date'], '%Y-%m-%d')
                month_num = date.month
                if month_num not in forecast_by_month:
                    forecast_by_month[month_num] = {'day_temps': [], 'night_temps': []}
                forecast_by_month[month_num]['day_temps'].append(day_data['day_temp'])
                forecast_by_month[month_num]['night_temps'].append(day_data['night_temp'])
        
        for month_num in range(1, 13):
            month_name = months[month_num - 1]
            month_data = {
                'month': month_name,
                'month_num': month_num,
                'years': {},
                'is_forecast': False,
                'source': 'Open-Meteo API'
            }
            
            for year in [2023, 2024, 2025, 2026]:
                year_monthly = yearly_data.get(year, [])
                month_info = next((m for m in year_monthly if m['month'] == month_num), None)

                # Check if this is a future month that needs forecast data
                is_future_month = (year == current_year and month_num > current_month)

                # For current year: skip months beyond forecast window
                if year == current_year and month_num > forecast_end_month:
                    month_data['years'][year] = None
                    continue
                
                # Mark as forecast if beyond current month
                if is_future_month:
                    month_data['is_forecast'] = True
                
                day_temp = None
                night_temp = None
                source = None
                
                # Get data from API result
                if month_info and month_info['avg_day_temp'] is not None:
                    day_temp = month_info['avg_day_temp']
                    night_temp = month_info['avg_night_temp']
                    source = month_info.get('source', 'Open-Meteo Archive')
                
                # Use forecast data for future months in current year
                elif is_future_month and month_num in forecast_by_month:
                    temps = forecast_by_month[month_num]
                    day_temp = round(sum(temps['day_temps']) / len(temps['day_temps']), 1)
                    night_temp = round(sum(temps['night_temps']) / len(temps['night_temps']), 1)
                    source = 'Open-Meteo Seasonal (ECMWF SEAS5)'
                
                if day_temp is not None:
                    # Calculate demand index based on night temp
                    demand = min(100, max(0, int((night_temp - 15) * 7)))
                    
                    # Calculate AC hours
                    ac_hours = max(0, min(24, round((night_temp - 18) * 2.5 + (day_temp - 30) * 0.5, 1)))
                    
                    month_data['years'][year] = {
                        'avg_day_temp': day_temp,
                        'avg_night_temp': night_temp,
                        'max_day_temp': round(day_temp + 2, 1),  # Estimate
                        'max_night_temp': round(night_temp + 2, 1),  # Estimate
                        'avg_demand': demand,
                        'avg_ac_hours': ac_hours,
                        'source': source
                    }
                else:
                    month_data['years'][year] = None
            
            # Calculate YoY changes
            if month_data['years'].get(2023) and month_data['years'].get(2024):
                month_data['yoy_2023_2024'] = {
                    'day_temp_change': round(month_data['years'][2024]['avg_day_temp'] - month_data['years'][2023]['avg_day_temp'], 1),
                    'night_temp_change': round(month_data['years'][2024]['avg_night_temp'] - month_data['years'][2023]['avg_night_temp'], 1),
                    'demand_change': round(month_data['years'][2024]['avg_demand'] - month_data['years'][2023]['avg_demand'], 1)
                }
            if month_data['years'].get(2024) and month_data['years'].get(2025):
                month_data['yoy_2024_2025'] = {
                    'day_temp_change': round(month_data['years'][2025]['avg_day_temp'] - month_data['years'][2024]['avg_day_temp'], 1),
                    'night_temp_change': round(month_data['years'][2025]['avg_night_temp'] - month_data['years'][2024]['avg_night_temp'], 1),
                    'demand_change': round(month_data['years'][2025]['avg_demand'] - month_data['years'][2024]['avg_demand'], 1)
                }
            if month_data['years'].get(2025) and month_data['years'].get(2026):
                month_data['yoy_2025_2026'] = {
                    'day_temp_change': round(month_data['years'][2026]['avg_day_temp'] - month_data['years'][2025]['avg_day_temp'], 1),
                    'night_temp_change': round(month_data['years'][2026]['avg_night_temp'] - month_data['years'][2025]['avg_night_temp'], 1),
                    'demand_change': round(month_data['years'][2026]['avg_demand'] - month_data['years'][2025]['avg_demand'], 1)
                }
            
            comparison_data.append(month_data)
        
        # Filter by specific month if requested
        if month_filter:
            try:
                month_num = int(month_filter)
                comparison_data = [m for m in comparison_data if m['month_num'] == month_num]
            except ValueError:
                pass
        
        return jsonify({
            'status': 'success',
            'data': comparison_data,
            'meta': {
                'city_filter': city_filter,
                'city_used': sample_city['name'],
                'month_filter': month_filter,
                'years_compared': [2023, 2024, 2025, 2026],
                'current_date': datetime.now().strftime('%Y-%m-%d'),
                'data_source': 'Open-Meteo API'
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/zones/demand-summary')
def get_zone_demand_summary():
    """
    Branch-wise (zone-wise) monthly demand summary — stacked view.
    Groups all 60 cities by demand_zone and returns monthly avg demand index per zone
    for the next 6 months (forecast) + last 6 months (historical).
    """
    try:
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year

        # Define zones with display colours
        zone_config = {
            'Extreme Heat Zone':       {'color': '#ef4444', 'short': 'Extreme Heat'},
            'Coastal Humid Zone':      {'color': '#3b82f6', 'short': 'Coastal Humid'},
            'Dry Heat Zone':           {'color': '#f97316', 'short': 'Dry Heat'},
            'Indo-Gangetic Heat Zone': {'color': '#eab308', 'short': 'Indo-Gangetic'},
            'Humid Heat Zone':         {'color': '#8b5cf6', 'short': 'Humid Heat'},
            'Moderate Plateau Zone':   {'color': '#22c55e', 'short': 'Moderate Plateau'},
            'North Plains Heat Zone':  {'color': '#06b6d4', 'short': 'North Plains'},
            'Manufacturing Hub':       {'color': '#ec4899', 'short': 'Manufacturing'},
        }

        # Group cities by zone
        zone_cities = {}
        for city in Config.CITIES:
            zone = city.get('demand_zone', 'Other')
            zone_cities.setdefault(zone, []).append(city)

        # Build 12-month window: 6 historical + current + 5 forecast
        months_window = []
        for offset in range(-5, 7):
            m = current_month + offset
            y = current_year
            while m < 1:
                m += 12
                y -= 1
            while m > 12:
                m -= 12
                y += 1
            months_window.append({'month': m, 'year': y,
                                   'label': datetime(y, m, 1).strftime('%b %Y'),
                                   'is_forecast': (y > current_year) or (y == current_year and m > current_month)})

        # Seasonal demand profile (night-temp driven, avg across India)
        seasonal_demand = {
            1: 28, 2: 32, 3: 48, 4: 70, 5: 85,
            6: 72, 7: 55, 8: 50, 9: 52, 10: 45, 11: 35, 12: 25
        }
        zone_demand_multipliers = {
            'Extreme Heat Zone': 1.35,
            'Coastal Humid Zone': 1.10,
            'Dry Heat Zone': 1.25,
            'Indo-Gangetic Heat Zone': 1.20,
            'Humid Heat Zone': 1.15,
            'Moderate Plateau Zone': 0.75,
            'North Plains Heat Zone': 1.10,
            'Manufacturing Hub': 1.05,
        }

        result = {
            'months': [m['label'] for m in months_window],
            'is_forecast': [m['is_forecast'] for m in months_window],
            'zones': []
        }

        for zone, zconf in zone_config.items():
            cities_in_zone = zone_cities.get(zone, [])
            if not cities_in_zone:
                continue
            multiplier = zone_demand_multipliers.get(zone, 1.0)
            city_count = len(cities_in_zone)
            demand_values = []
            for mw in months_window:
                base = seasonal_demand.get(mw['month'], 40)
                # Each city in zone contributes to zone-total demand
                zone_demand = round(min(100, base * multiplier), 1)
                demand_values.append(zone_demand)
            result['zones'].append({
                'zone': zone,
                'short': zconf['short'],
                'color': zconf['color'],
                'city_count': city_count,
                'demand': demand_values
            })

        return jsonify({'status': 'success', 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/forecast/model-mix')
def get_model_mix():
    """
    AC product model-mix forecast.
    Returns percentage split across model categories for the next 6 months
    based on demand intensity. Higher demand → higher-capacity/inverter mix.
    """
    try:
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year

        # Month labels for next 6 months
        months = []
        for offset in range(6):
            m = current_month + offset
            y = current_year
            while m > 12:
                m -= 12
                y += 1
            months.append(datetime(y, m, 1).strftime('%b %Y'))

        # Seasonal base demand index (1-100) — drives model mix
        seasonal_demand = {1: 28, 2: 32, 3: 48, 4: 70, 5: 85, 6: 72,
                           7: 55, 8: 50, 9: 52, 10: 45, 11: 35, 12: 25}

        model_categories = [
            {'id': 'split_1t_inv',  'label': '1T Inverter Split',    'color': '#3b82f6'},
            {'id': 'split_15t_inv', 'label': '1.5T Inverter Split',  'color': '#10b981'},
            {'id': 'split_2t_inv',  'label': '2T Inverter Split',    'color': '#f59e0b'},
            {'id': 'split_nonInv',  'label': 'Non-Inverter Split',   'color': '#ef4444'},
            {'id': 'window_ac',     'label': 'Window AC',            'color': '#8b5cf6'},
            {'id': 'cassette_com',  'label': 'Cassette/Commercial',  'color': '#ec4899'},
        ]

        # Model mix shifts with demand: low demand → budget/non-inv; high demand → premium inverter + 2T
        def compute_mix(demand_idx):
            d = demand_idx / 100.0
            split_1t   = round(max(5,  28 - d * 10), 1)
            split_15t  = round(min(40, 22 + d * 18), 1)
            split_2t   = round(min(25,  5 + d * 20), 1)
            non_inv    = round(max(5,  25 - d * 20), 1)
            window     = round(max(3,  12 - d *  8), 1)
            cassette   = round(max(2,   8 - d *  3), 1)
            total = split_1t + split_15t + split_2t + non_inv + window + cassette
            # Normalise to 100
            factor = 100 / total
            return [round(split_1t * factor, 1), round(split_15t * factor, 1),
                    round(split_2t * factor, 1),  round(non_inv * factor, 1),
                    round(window * factor, 1),     round(cassette * factor, 1)]

        result = {'months': months, 'models': []}
        monthly_mixes = []
        for offset in range(6):
            m = (current_month + offset - 1) % 12 + 1
            monthly_mixes.append(compute_mix(seasonal_demand.get(m, 40)))

        for idx, model in enumerate(model_categories):
            result['models'].append({
                'label': model['label'],
                'color': model['color'],
                'values': [monthly_mixes[mo][idx] for mo in range(6)]
            })

        # Overall mix (avg across 6 months)
        for model_data in result['models']:
            model_data['avg'] = round(sum(model_data['values']) / len(model_data['values']), 1)

        return jsonify({'status': 'success', 'data': result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


def generate_strategic_recommendations(cities_weather, forecast_by_city, historical_by_city, current_month):
    """Generate strategic business recommendations"""
    recommendations = []
    
    # High priority: Peak season preparation
    hot_cities = [c for c in cities_weather if (c.get('night_temp') or 0) >= 22]
    if len(hot_cities) >= 3 or current_month in [3, 4, 5]:
        recommendations.append({
            'priority': 'high',
            'icon': 'fire',
            'title': 'Peak Season Approaching',
            'description': f'{len(hot_cities)} cities showing elevated night temperatures. Accelerate inventory positioning for AC and cooling products.',
            'metrics': [
                {'value': f'{len(hot_cities)}', 'label': 'Hot Markets'},
                {'value': 'Next 6 weeks', 'label': 'Action Window'}
            ]
        })
    
    # Identify fastest heating markets
    heating_markets = []
    for city in Config.CITIES:
        forecast = forecast_by_city.get(city['id'], [])
        if len(forecast) >= 30:
            first_week = sum(d['night_temp'] for d in forecast[:7]) / 7
            fourth_week = sum(d['night_temp'] for d in forecast[21:28]) / 7 if len(forecast) >= 28 else first_week
            if fourth_week - first_week > 2:
                heating_markets.append(city['name'])
    
    if heating_markets:
        recommendations.append({
            'priority': 'high',
            'icon': 'chart-line',
            'title': 'Rapid Temperature Rise Detected',
            'description': f'{", ".join(heating_markets[:3])} showing rapid heating trend. Priority markets for immediate stock deployment.',
            'metrics': [
                {'value': f'+2-4°C', 'label': 'Expected Rise'},
                {'value': '4 weeks', 'label': 'Timeline'}
            ]
        })
    
    # Night temperature opportunity
    night_opportunity = [c for c in cities_weather if 20 <= (c.get('night_temp') or 0) < 24]
    if night_opportunity:
        recommendations.append({
            'priority': 'medium',
            'icon': 'moon',
            'title': 'Night Temperature Sweet Spot',
            'description': f'{len(night_opportunity)} markets in 20-24°C night range. Optimal for premium AC positioning - consumers actively seeking comfort.',
            'metrics': [
                {'value': '12-16 hrs', 'label': 'AC Usage'},
                {'value': 'Premium', 'label': 'Segment Focus'}
            ]
        })
    
    # Regional strategy based on data
    # Identify hottest South Indian cities by night temp
    south_hot_cities = []
    for c in cities_weather:
        night_temp = c.get('night_temp') or 0
        if night_temp >= 22:
            south_hot_cities.append({'name': c['city_name'], 'night_temp': round(night_temp, 1)})
    south_hot_cities.sort(key=lambda x: x['night_temp'], reverse=True)
    
    if south_hot_cities:
        city_list = ', '.join([f"{c['name']} ({c['night_temp']}°C)" for c in south_hot_cities[:3]])
        recommendations.append({
            'priority': 'medium',
            'icon': 'map-marker-alt',
            'title': f'{len(south_hot_cities)} Cities Showing High Night Temps',
            'description': f'{city_list} — prioritize inventory allocation to these markets.',
            'metrics': [
                {'value': f"{south_hot_cities[0]['night_temp']}°C", 'label': f"Hottest: {south_hot_cities[0]['name']}"},
                {'value': f'{len(south_hot_cities)}', 'label': 'Hot Markets'}
            ]
        })
    
    # Inventory timing recommendation
    recommendations.append({
        'priority': 'low',
        'icon': 'warehouse',
        'title': 'Inventory Timing Optimal',
        'description': 'Based on 4-month forecast, current period ideal for channel loading. Distributor stock-up recommended before March price increases.',
        'metrics': [
            {'value': '2-3 weeks', 'label': 'Lead Time'},
            {'value': 'Feb-Mar', 'label': 'Optimal Window'}
        ]
    })
    
    # Promotional timing
    if current_month in [2, 3]:
        recommendations.append({
            'priority': 'low',
            'icon': 'bullhorn',
            'title': 'Pre-Season Promotion Window',
            'description': 'Consumer awareness campaigns effective now. Digital marketing ROI highest 4-6 weeks before peak.',
            'metrics': [
                {'value': '4-6 weeks', 'label': 'Before Peak'},
                {'value': '2x', 'label': 'Conversion Rate'}
            ]
        })
    
    return recommendations[:6]  # Return top 6


def generate_monthly_forecast_breakdown(forecast_by_city, current_month):
    """Generate monthly forecast breakdown with demand levels"""
    months_data = []
    month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    for i in range(4):  # Next 4 months
        target_month = current_month + i
        if target_month > 12:
            break
        
        all_day_temps = []
        all_night_temps = []
        
        for city_id, forecast in forecast_by_city.items():
            month_data = [d for d in forecast if datetime.strptime(d['date'], '%Y-%m-%d').month == target_month]
            for d in month_data:
                all_day_temps.append(d['day_temp'])
                all_night_temps.append(d['night_temp'])
        
        if all_day_temps and all_night_temps:
            avg_day = sum(all_day_temps) / len(all_day_temps)
            avg_night = sum(all_night_temps) / len(all_night_temps)
            demand = min(100, max(0, int((avg_night - 15) * 7)))
            
            # Determine demand level
            if demand >= 80:
                level = 'extreme'
            elif demand >= 60:
                level = 'high'
            elif demand >= 40:
                level = 'moderate'
            else:
                level = 'normal'
            
            months_data.append({
                'month': month_names[target_month - 1],
                'month_num': target_month,
                'avg_day': round(avg_day, 1),
                'avg_night': round(avg_night, 1),
                'demand': demand,
                'level': level
            })
    
    return months_data


def generate_action_checklist(days_to_peak, demand_potential, city_scores, current_month):
    """Generate prioritized action checklist"""
    actions = []
    
    # Urgency based on days to peak
    if days_to_peak <= 30:
        actions.append({
            'urgency': 'urgent',
            'title': 'Emergency Stock Deployment',
            'detail': 'Peak season imminent. Prioritize logistics and retailer stock levels.',
            'timeline': 'This week'
        })
    elif days_to_peak <= 60:
        actions.append({
            'urgency': 'urgent',
            'title': 'Accelerate Channel Loading',
            'detail': 'Push inventory to distributors in top 5 priority markets.',
            'timeline': 'Next 2 weeks'
        })
    
    # High demand potential actions
    if demand_potential >= 70:
        actions.append({
            'urgency': 'important',
            'title': 'Premium Product Push',
            'detail': 'High-margin inverter AC models should lead portfolio mix.',
            'timeline': 'Ongoing'
        })
    
    # Top market specific
    if city_scores and city_scores[0]['score'] >= 70:
        top = city_scores[0]
        actions.append({
            'urgency': 'important',
            'title': f'Focus on {top["city"]}',
            'detail': f'Highest potential market with {top["night_temp"]}°C nights. Increase retail presence.',
            'timeline': 'Next 4 weeks'
        })
    
    # Seasonal actions
    if current_month == 2:
        actions.append({
            'urgency': 'important',
            'title': 'Trade Scheme Launch',
            'detail': 'Announce dealer incentives for pre-season orders.',
            'timeline': 'Feb 15-28'
        })
    
    if current_month in [2, 3]:
        actions.append({
            'urgency': 'normal',
            'title': 'Consumer Financing Tie-ups',
            'detail': 'Finalize EMI partnerships with banks for peak season.',
            'timeline': 'Before March end'
        })
    
    actions.append({
        'urgency': 'normal',
        'title': 'Service Network Readiness',
        'detail': 'Ensure installation capacity scaled up for anticipated demand.',
        'timeline': 'Next 6 weeks'
    })
    
    actions.append({
        'urgency': 'normal',
        'title': 'Digital Marketing Campaign',
        'detail': 'Launch awareness campaigns targeting AC research keywords.',
        'timeline': 'Ongoing'
    })
    
    return actions[:8]


def generate_city_insights(cities_weather, forecast_by_city, historical_by_city):
    """Generate insights for each city"""
    city_insights = []
    
    for city in cities_weather:
        city_id = city['city_id']
        forecast = forecast_by_city.get(city_id, [])

        # Calculate metrics (use 'or' to treat None as missing)
        current_night = city.get('night_temp') or 20
        current_day = city.get('day_temp') or 32

        # Future projection
        future_nights = [d['night_temp'] for d in forecast[:60] if d.get('night_temp') is not None]
        future_days = [d['day_temp'] for d in forecast[:60] if d.get('day_temp') is not None]

        avg_future_night = sum(future_nights) / len(future_nights) if future_nights else current_night
        avg_future_day = sum(future_days) / len(future_days) if future_days else current_day

        # Calculate demand score
        demand_score = min(100, max(0, int((avg_future_night - 15) * 7)))

        # Determine priority
        if demand_score >= 70:
            priority = 'high'
            badge = 'High Priority'
        elif demand_score >= 50:
            priority = 'medium'
            badge = 'Medium Priority'
        else:
            priority = 'low'
            badge = 'Monitor'

        # Generate recommendation
        if avg_future_night >= 24:
            recommendation = f"<strong>Peak demand territory.</strong> Night temps above 24°C drive 14-16 hr AC usage. Prioritize premium models and installation capacity."
        elif avg_future_night >= 22:
            recommendation = f"<strong>Strong demand zone.</strong> Current {current_night}°C nights rising to {avg_future_night:.1f}°C. Accelerate inventory deployment."
        elif avg_future_night >= 20:
            recommendation = f"<strong>Building demand.</strong> Position stock now for upcoming surge. Focus on mid-range segment."
        else:
            recommendation = f"<strong>Moderate demand.</strong> Maintain baseline inventory. Monitor for temperature changes."

        # Demand zone from config
        city_config = next((c for c in Config.CITIES if c['id'] == city_id), {})

        city_insights.append({
            'city': city['city_name'],
            'city_id': city_id,
            'priority': priority,
            'badge': badge,
            'current_day': round(current_day, 1),
            'current_night': round(current_night, 1),
            'forecast_night': round(avg_future_night, 1),
            'demand_score': demand_score,
            'recommendation': recommendation,
            'demand_zone': city_config.get('demand_zone', ''),
            'zone_icon': city_config.get('zone_icon', '📍')
        })
    
    # Sort by demand score
    city_insights.sort(key=lambda x: x['demand_score'], reverse=True)
    
    return city_insights


# ═══════════════════════════════════════════════════════════════════════════
# NEW API ENDPOINTS (Boss's 6-Point Directive)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/api/advanced-weather')
def get_advanced_weather():
    """
    Advanced weather parameters: wet bulb, heat wave, monsoon, consecutive hot days.
    Query params: city (optional, default: all)
    """
    try:
        city_filter = request.args.get('city', 'all')
        cities_weather = get_cached_weather()
        
        if city_filter != 'all':
            cities_weather = [c for c in cities_weather if c.get('city_id') == city_filter]
        
        result = []
        for city in cities_weather:
            city_id = city.get('city_id', '')
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            humidity = city.get('humidity') or 60
            
            # Wet bulb
            wet_bulb = data_processor.calculate_wet_bulb(day_temp, humidity)
            
            # Consecutive hot days from forecast
            forecast = get_cached_forecast(city_id, days=16)
            heatwave = data_processor.detect_consecutive_hot_days(forecast)
            
            # DSB zone
            demand_idx = data_processor.calculate_demand_index(day_temp, night_temp, humidity)
            dsb = data_processor.get_dsb_zone(demand_idx)
            
            # City config
            city_config = next((c for c in Config.CITIES if c['id'] == city_id), {})
            
            result.append({
                'city_id': city_id,
                'city_name': city.get('city_name', ''),
                'demand_zone': city_config.get('demand_zone', ''),
                'zone_icon': city_config.get('zone_icon', '📍'),
                'zone_traits': city_config.get('zone_traits', ''),
                'wet_bulb': wet_bulb,
                'heatwave': heatwave,
                'dsb_zone': dsb,
                'demand_index': round(demand_idx, 1),
                'day_temp': day_temp,
                'night_temp': night_temp,
                'humidity': humidity
            })
        
        # Monsoon status (same for all cities in region)
        monsoon = data_processor.get_monsoon_status()
        
        return jsonify({
            'status': 'success',
            'data': {
                'cities': result,
                'monsoon': monsoon
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500





@app.route('/api/alerts/ack', methods=['POST'], endpoint='alerts_ack')
def acknowledge_alerts():
    """Acknowledge alerts by id. Accepts JSON: {'ids': ['city_...'] }"""
    try:
        body = request.get_json() or {}
        ids = body.get('ids', [])
        if not isinstance(ids, list):
            return jsonify({'status': 'error', 'message': 'ids must be a list'}), 400
        for _id in ids:
            cache.setdefault('alerts_ack', set()).add(_id)
        # update cached alerts' acknowledged flags
        for a in cache.get('alerts_data', []):
            if a.get('id') in ids:
                a['acknowledged'] = True
        return jsonify({'status': 'success', 'acknowledged': ids})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/generate-checklist', endpoint='generate_checklist_api')
def generate_checklist():
    """Generate an actionable checklist for a city based on current conditions. Query: city=<city_id> or city=all"""
    try:
        city_id = request.args.get('city')
        if not city_id or city_id == 'all':
            return jsonify({'status': 'error', 'message': 'Please specify a city parameter'}), 400
        # Find city weather
        cities = get_cached_weather()
        city = next((c for c in cities if c.get('city_id') == city_id), None)
        if not city:
            return jsonify({'status': 'error', 'message': 'City not found'}), 404

        # Use alert engine to create recommendation steps
        alert = alert_engine.analyze_temperature(
            _safe_temp(city, 'day_temp', 32),
            _safe_temp(city, 'night_temp', 20),
            city.get('city_name'),
            city_id
        )
        rec = alert.get('recommendation', {})
        checklist = {
            'city': city.get('city_name'),
            'city_id': city_id,
            'demand_index': alert.get('demand_index'),
            'action': rec.get('action'),
            'priority': rec.get('priority'),
            'steps': rec.get('steps', [])
        }
        return jsonify({'status': 'success', 'data': {'checklist': checklist}})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/demand-correlation')
def get_demand_correlation():
    """
    Demand correlation layer: overlay historical weather vs simulated sales data.
    Query params: city (city_id, default: chennai), days (int, default: 90)
    """
    try:
        city_id = request.args.get('city', 'chennai')
        days = int(request.args.get('days', 90))
        
        # Get historical weather data
        historical = weather_service.get_historical_data(city_id, days=days)
        
        if not historical:
            return jsonify({'status': 'error', 'message': 'No historical data available'}), 404
        
        # Calculate correlation
        correlation = data_processor.calculate_demand_correlation(historical, city_id)
        
        city_config = next((c for c in Config.CITIES if c['id'] == city_id), {})
        
        return jsonify({
            'status': 'success',
            'data': {
                'city_id': city_id,
                'city_name': city_config.get('name', city_id),
                'demand_zone': city_config.get('demand_zone', ''),
                'days_analyzed': len(historical),
                'correlation': correlation
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/service-predictions')
def get_service_predictions():
    """
    Service demand predictions: compressor failures, gas refills, warranty claims.
    Returns predictions for all cities or a specific city.
    Query params: city (optional, default: all)
    """
    try:
        city_filter = request.args.get('city', 'all')
        cities_weather = get_cached_weather()
        
        if city_filter != 'all':
            cities_weather = [c for c in cities_weather if c.get('city_id') == city_filter]
        
        predictions = []
        for city in cities_weather:
            city_id = city.get('city_id', '')
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            humidity = city.get('humidity') or 60
            
            pred = data_processor.predict_service_demand(city_id, day_temp, night_temp, humidity)
            pred['city_id'] = city_id
            pred['city_name'] = city.get('city_name', '')
            
            city_config = next((c for c in Config.CITIES if c['id'] == city_id), {})
            pred['demand_zone'] = city_config.get('demand_zone', '')
            
            # Dynamic Zone Icon based on load
            load = pred.get('overall_service_load', 0)
            if load > 75:
                pred['zone_icon'] = "🔥"
            elif load > 50:
                pred['zone_icon'] = "⚠️"
            else:
                pred['zone_icon'] = "❄️"
            
            predictions.append(pred)

        # Sort by overall service load
        predictions.sort(key=lambda x: x.get('overall_service_load', 0), reverse=True)
        
        return jsonify({
            'status': 'success',
            'data': predictions
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/refresh-status')
def get_refresh_status():
    """
    Get refresh cadence status: daily weather, weekly demand forecast, monthly accuracy validation.
    """
    try:
        refresh = data_processor.get_refresh_status()
        return jsonify({
            'status': 'success',
            'data': refresh
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/dsb-overview')
def get_dsb_overview():
    """
    DSB methodology overview: Green/Amber/Red zones for all cities.
    """
    try:
        cities_weather = get_cached_weather()
        
        zones = {'green': [], 'amber': [], 'red': []}
        
        for city in cities_weather:
            demand_idx = city.get('demand_index', 0)
            dsb = data_processor.get_dsb_zone(demand_idx)
            city_config = next((c for c in Config.CITIES if c['id'] == city.get('city_id')), {})
            
            entry = {
                'city_id': city.get('city_id', ''),
                'city_name': city.get('city_name', ''),
                'demand_index': round(demand_idx, 1),
                'demand_zone': city_config.get('demand_zone', ''),
                'zone_icon': city_config.get('zone_icon', '📍'),
                'day_temp': city.get('day_temp'),
                'night_temp': city.get('night_temp'),
                'ac_hours': round(city.get('ac_hours', 0), 1),
                'humidity': round(city.get('humidity', 0), 1),
                'wet_bulb': round(city['wet_bulb']['value'], 1) if isinstance(city.get('wet_bulb'), dict) else (round(city['wet_bulb'], 1) if city.get('wet_bulb') is not None else None),
                'dsb': dsb,
                'action': dsb.get('action', 'Monitor'),
            }
            zones[dsb['zone']].append(entry)
        
        return jsonify({
            'status': 'success',
            'data': {
                'zones': zones,
                'summary': {
                    'green_count': len(zones['green']),
                    'amber_count': len(zones['amber']),
                    'red_count': len(zones['red']),
                    'total': len(cities_weather)
                },
                'methodology': {
                    'green': Config.DSB_GREEN,
                    'amber': Config.DSB_AMBER,
                    'red': Config.DSB_RED
                }
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/demand-intel-combined')
def get_demand_intel_combined():
    """
    Combined endpoint for Demand Intel page — returns dsb-overview, demand-prediction,
    energy-estimates, and service-predictions in a single response.
    Avoids 4 separate API calls that each redundantly call get_cached_weather().
    """
    try:
        cities_weather = get_cached_weather()
        
        # Ensure all cities have required fields (normalize compact format if needed)
        for city in cities_weather:
            # Handle compact format
            if 'n' in city and 'city_name' not in city:
                city['city_name'] = city.get('n')
                city['city_id'] = city.get('id')
                city['day_temp'] = city.get('dt')
                city['night_temp'] = city.get('nt')
                city['temperature'] = city.get('t')
                city['humidity'] = city.get('h')
                city['demand_index'] = city.get('di')
            # Ensure derived fields exist
            if city.get('ac_hours') is None:
                city['ac_hours'] = data_processor.calculate_ac_hours(
                    city.get('day_temp'), city.get('night_temp')
                )
            if city.get('wet_bulb') is None or isinstance(city.get('wet_bulb'), dict):
                wb = data_processor.calculate_wet_bulb(city.get('day_temp'), city.get('humidity'))
                city['wet_bulb'] = wb.get('value') if isinstance(wb, dict) and wb.get('value') is not None else (wb if isinstance(wb, (int, float)) else None)

        # 1. DSB Overview
        zones = {'green': [], 'amber': [], 'red': []}
        for city in cities_weather:
            demand_idx = city.get('demand_index', 0)
            dsb = data_processor.get_dsb_zone(demand_idx)
            city_config = next((c for c in Config.CITIES if c['id'] == city.get('city_id')), {})
            
            # Get wet_bulb value (handle both dict and numeric formats)
            wet_bulb_val = city.get('wet_bulb')
            if isinstance(wet_bulb_val, dict):
                wet_bulb_val = wet_bulb_val.get('value')
            
            entry = {
                'city_id': city.get('city_id', ''),
                'city_name': city.get('city_name', ''),
                'demand_index': round(demand_idx, 1) if demand_idx else 0,
                'demand_zone': city_config.get('demand_zone', ''),
                'zone_icon': city_config.get('zone_icon', ''),
                'day_temp': city.get('day_temp'),
                'night_temp': city.get('night_temp'),
                'ac_hours': round(city.get('ac_hours', 0), 1) if city.get('ac_hours') is not None else None,
                'humidity': round(city.get('humidity', 0), 1) if city.get('humidity') is not None else None,
                'wet_bulb': round(wet_bulb_val, 1) if wet_bulb_val is not None else None,
                'dsb': dsb,
                'action': dsb.get('action', 'Monitor'),
            }
            zones[dsb['zone']].append(entry)
        dsb_data = {
            'zones': zones,
            'summary': {
                'green_count': len(zones['green']),
                'amber_count': len(zones['amber']),
                'red_count': len(zones['red']),
                'total': len(cities_weather)
            },
            'methodology': {
                'green': Config.DSB_GREEN,
                'amber': Config.DSB_AMBER,
                'red': Config.DSB_RED
            }
        }

        # 2. Demand Predictions
        predictions = []
        for city in cities_weather:
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            humidity = city.get('humidity') or 60
            night_score = min(100, max(0, (night_temp - 15) * 10))
            day_score = min(100, max(0, (day_temp - 25) * 7.5))
            humidity_score = min(100, max(0, (humidity - 40) * 1.5))
            demand_score = round(night_score * 0.6 + day_score * 0.25 + humidity_score * 0.15, 1)
            if demand_score >= 80:
                demand_level, recommendation = 'Critical', 'Maximum inventory allocation'
            elif demand_score >= 60:
                demand_level, recommendation = 'High', 'Increase stock levels'
            elif demand_score >= 40:
                demand_level, recommendation = 'Moderate', 'Maintain current levels'
            else:
                demand_level, recommendation = 'Low', 'Reduce inventory focus'
            predictions.append({
                'city': city['city_name'], 'city_id': city['city_id'],
                'demand_score': demand_score, 'demand_level': demand_level,
                'confidence': 'High' if city.get('source', '') != 'Simulated' else 'Medium',
                'recommendation': recommendation,
                'factors': {
                    'night_temp_contribution': round(night_score * 0.6, 1),
                    'day_temp_contribution': round(day_score * 0.25, 1),
                    'humidity_contribution': round(humidity_score * 0.15, 1)
                }
            })
        predictions.sort(key=lambda x: x['demand_score'], reverse=True)

        # 3. Energy Estimates
        estimates = []
        for city in cities_weather:
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            day_ac = max(0, min(12, (day_temp - 28) * 1.2))
            night_ac = max(0, min(12, (night_temp - 20) * 3))
            total_ac = round(day_ac + night_ac, 1)
            daily_kwh = round(total_ac * 1.5, 1)
            monthly_kwh = round(daily_kwh * 30, 1)
            monthly_cost = round(monthly_kwh * 6.5, 0)
            estimates.append({
                'city': city['city_name'], 'city_id': city['city_id'],
                'ac_hours_day': round(day_ac, 1), 'ac_hours_night': round(night_ac, 1),
                'total_ac_hours': total_ac, 'daily_kwh': daily_kwh,
                'monthly_kwh': monthly_kwh, 'estimated_monthly_cost': f'\u20b9{monthly_cost:,.0f}'
            })

        # 4. Service Predictions
        svc_predictions = []
        for city in cities_weather:
            city_id = city.get('city_id', '')
            day_temp = _safe_temp(city, 'day_temp', 32)
            night_temp = _safe_temp(city, 'night_temp', 20)
            humidity = city.get('humidity') or 60
            pred = data_processor.predict_service_demand(city_id, day_temp, night_temp, humidity)
            pred['city_id'] = city_id
            pred['city_name'] = city.get('city_name', '')
            city_config = next((c for c in Config.CITIES if c['id'] == city_id), {})
            pred['demand_zone'] = city_config.get('demand_zone', '')
            load = pred.get('overall_service_load', 0)
            pred['zone_icon'] = "🔥" if load > 75 else ("⚠️" if load > 50 else "❄️")
            svc_predictions.append(pred)
        svc_predictions.sort(key=lambda x: x.get('overall_service_load', 0), reverse=True)

        return jsonify({
            'status': 'success',
            'data': {
                'dsb': dsb_data,
                'demand': predictions,
                'energy': estimates,
                'service': svc_predictions
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cache-stats')
def get_cache_stats():
    """Expose cache monitoring stats for debugging and observability."""
    avg_latency = (cache_stats['total_api_latency'] / cache_stats['api_call_count']
                   if cache_stats['api_call_count'] > 0 else 0)
    return jsonify({
        'status': 'success',
        'data': {
            'cache_hits': cache_stats['cache_hits'],
            'cache_misses': cache_stats['cache_misses'],
            'hit_rate': round(cache_stats['cache_hits'] / max(1, cache_stats['cache_hits'] + cache_stats['cache_misses']) * 100, 1),
            'api_calls': cache_stats['api_calls'],
            'api_errors': cache_stats['api_errors'],
            'avg_latency_seconds': round(avg_latency, 3),
            'last_refresh': cache_stats['last_refresh'],
            'weather_cache_age': round(time.time() - cache['weather_timestamp'], 1) if cache['weather_timestamp'] else None,
            'alerts_cache_age': round(time.time() - cache['alerts_timestamp'], 1) if cache['alerts_timestamp'] else None,
            'weather_stale': cache.get('weather_data_stale', False),
            'sse_clients': cache_stats.get('sse_clients', 0)
        }
    })


@app.route('/api/stream')
def sse_stream():
    """Server-Sent Events endpoint for real-time dashboard updates."""
    def generate():
        q = queue.Queue(maxsize=50)
        with sse_clients_lock:
            sse_clients.append(q)
            cache_stats['sse_clients'] = len(sse_clients)
        try:
            yield "event: connected\ndata: {}\n\n"
            while True:
                try:
                    message = q.get(timeout=30)
                    yield message
                except queue.Empty:
                    yield ": keepalive\n\n"
        except GeneratorExit:
            pass
        finally:
            with sse_clients_lock:
                if q in sse_clients:
                    sse_clients.remove(q)
                cache_stats['sse_clients'] = len(sse_clients)

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


def warm_cache():
    """Pre-warm caches in background so first user request is instant"""
    try:
        time.sleep(1)  # Wait for server to be fully ready

        print("[Cache] Pre-warming weather cache...")
        weather = get_cached_weather()
        print(f"[Cache] Weather ready ({len(weather)} cities)")

        # DISABLED: Forecast pre-warming to avoid rate limits on startup
        # Forecasts will load on-demand when users request them
        print("[Cache] Forecast pre-warming DISABLED (loads on-demand to avoid rate limits)")

        # DISABLED: Historical data pre-warming to avoid rate limits on startup
        # Historical data will load on-demand when users request it
        print("[Cache] Historical pre-warming DISABLED (loads on-demand to avoid rate limits)")

        print("[Cache] Cache warming completed (current weather only)")
    except Exception as e:
        print(f"[Cache] Warning: Pre-warming failed: {e}")


# ---- Supabase Database API Endpoints ----

@app.route('/api/supabase/weather-history')
def supabase_weather_history():
    """Get historical weather data from Supabase. Query: city=<city_id>, limit=<n>"""
    try:
        city_id = request.args.get('city', 'chennai')
        limit = int(request.args.get('limit', 24))
        data = supabase_handler.get_recent_logs(city_id, limit=limit)
        return jsonify({
            'status': 'success',
            'supabase_enabled': supabase_handler.enabled,
            'data': data,
            'count': len(data)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/supabase/alerts')
def supabase_alerts():
    """Get alerts from Supabase. Query: active_only=true|false"""
    try:
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        if active_only:
            data = supabase_handler.get_active_alerts()
        else:
            # Fetch all alerts (up to 100)
            if supabase_handler.enabled:
                response = supabase_handler.client.table("alerts")\
                    .select("*")\
                    .order("created_at", desc=True)\
                    .limit(100)\
                    .execute()
                data = response.data
            else:
                data = []
        return jsonify({
            'status': 'success',
            'supabase_enabled': supabase_handler.enabled,
            'data': data,
            'count': len(data)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/supabase/alerts/ack', methods=['POST'])
def supabase_ack_alert():
    """Acknowledge an alert in Supabase by UUID. Body: { "id": "<uuid>" }"""
    try:
        body = request.get_json() or {}
        alert_id = body.get('id')
        if not alert_id:
            return jsonify({'status': 'error', 'message': 'Missing alert id'}), 400
        result = supabase_handler.acknowledge_alert(alert_id)
        if result is not None:
            return jsonify({'status': 'success', 'acknowledged': alert_id})
        else:
            return jsonify({'status': 'error', 'message': 'Failed to acknowledge or Supabase disabled'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/supabase/status')
def supabase_status():
    """Check Supabase connection health."""
    status = supabase_handler.get_connection_status()
    return jsonify({
        'status': 'success',
        'supabase_enabled': supabase_handler.enabled,
        'connection': status
    })







# ═══════════════════════════════════════════════════════════════════════════
# APScheduler Background Jobs — proactive cache refresh
# ═══════════════════════════════════════════════════════════════════════════

def scheduled_weather_refresh():
    """Refresh weather data for all cities (runs every 3 hours per client requirement)."""
    if not _weather_fetch_lock.acquire(blocking=False):
        print("[Scheduler] Weather fetch already in progress, skipping this cycle")
        return
    try:
        print("[Scheduler] Refreshing weather data...")
        t0 = time.time()
        fresh_data = weather_service.get_all_cities_current()
        cache_stats['api_calls'] += 1
        cache_stats['total_api_latency'] += time.time() - t0
        cache_stats['api_call_count'] += 1

        if not fresh_data:
            print("[Scheduler] Weather refresh returned no data")
            return

        _process_weather_cities(fresh_data)
        processed_data = _ensure_all_cities_present(fresh_data)
        cache['weather_data'] = processed_data
        cache['weather_timestamp'] = time.time()
        cache['weather_data_stale'] = False
        cache['weather_data_age'] = 0
        file_cache.set('weather_all_cities', processed_data)
        cache_stats['last_refresh'] = datetime.now().isoformat()

        if supabase_handler.enabled and Config.PERSIST_WEATHER_LOGS:
            try:
                supabase_handler.save_weather_logs_batch(fresh_data)
            except Exception as e:
                print(f"[Supabase] Weather batch persist error: {e}")

        elapsed = time.time() - t0
        print(f"[Scheduler] Weather refreshed: {len(processed_data)} cities in {elapsed:.1f}s")
        notify_sse_clients('weather_update')
    except Exception as e:
        cache_stats['api_errors'] += 1
        print(f"[Scheduler] Weather refresh failed: {e}")
    finally:
        _weather_fetch_lock.release()


def scheduled_alerts_refresh():
    """Refresh alerts (runs every 3 hours per client requirement)."""
    try:
        refresh_alerts(force=True)
        notify_sse_clients('alerts_update')
    except Exception as e:
        print(f"[Scheduler] Alerts refresh failed: {e}")


def scheduled_forecast_refresh():
    """Refresh forecasts for key cities (runs every 3 hours per client requirement).
    Uses get_cached_forecast to avoid duplicate API calls if data is already fresh.
    """
    key_cities = ['mumbai', 'delhi', 'chennai', 'bangalore', 'hyderabad',
                  'kolkata', 'ahmedabad', 'pune', 'lucknow', 'kochi']
    refreshed = 0
    for city_id in key_cities:
        try:
            cache_key = f"{city_id}_120"
            now = time.time()
            # Skip if in-memory cache is still fresh (< 25 min old, slightly under 30 min interval)
            if (cache_key in cache['forecast_data'] and
                cache_key in cache['forecast_timestamp'] and
                (now - cache['forecast_timestamp'].get(cache_key, 0)) < FORECAST_CACHE_TTL - 300):
                continue
            data = weather_service.get_forecast(city_id, days=120)
            cache['forecast_data'][cache_key] = data
            cache['forecast_timestamp'][cache_key] = now
            file_cache.set(f'forecast_{cache_key}', data)
            refreshed += 1
            time.sleep(1)
        except Exception as e:
            cache_stats['api_errors'] += 1
            print(f"[Scheduler] Forecast refresh failed for {city_id}: {e}")
    print(f"[Scheduler] Forecast refresh: {refreshed}/{len(key_cities)} cities updated")


# Initialize and start the scheduler
# Guard against Flask debug reloader spawning duplicate schedulers
import os as _os
if not app.debug or _os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    scheduler = BackgroundScheduler(daemon=True)
    # Client requirement: pull data at most once every 3 hours.
    # Startup preload handles the immediate first fetch; scheduler takes over from there.
    scheduler.add_job(
        scheduled_weather_refresh,
        IntervalTrigger(hours=3),
        id='weather_refresh',
        name='Refresh weather data',
        next_run_time=datetime.now() + timedelta(hours=3)
    )
    scheduler.add_job(
        scheduled_alerts_refresh,
        IntervalTrigger(hours=3),
        id='alerts_refresh',
        name='Refresh alerts',
        next_run_time=datetime.now() + timedelta(hours=3)
    )
    scheduler.add_job(
        scheduled_forecast_refresh,
        IntervalTrigger(hours=3),
        id='forecast_refresh',
        name='Refresh forecasts for key cities',
        next_run_time=datetime.now() + timedelta(hours=3)
    )
    scheduler.start()
    print("[Scheduler] Background scheduler started (weather=3h, alerts=3h, forecast=3h) — data pulled once every 3 hours")

    # Preload weather cache on startup so first user request is instant
    def _startup_preload():
        """Background preload: fetch weather immediately so users never wait"""
        # No sleep — acquire lock immediately to win the race against the first user request.
        # If a user request comes before us, it will do its own fetch and we'll skip.
        if not _weather_fetch_lock.acquire(blocking=False):
            return  # Another thread already fetching
        try:
            # Check if file cache already has data
            file_data = file_cache.get('weather_all_cities')
            if file_data:
                _process_weather_cities(file_data)
                processed = _ensure_all_cities_present(file_data)
                cache['weather_data'] = processed
                cache['weather_timestamp'] = time.time()
                print(f"[Startup] Preloaded {len(processed)} cities from file cache")
                return

            # No file cache — fetch from API
            print("[Startup] No file cache, fetching weather from API...")
            t0 = time.time()
            fresh = weather_service.get_all_cities_current()
            if fresh:
                _process_weather_cities(fresh)
                processed = _ensure_all_cities_present(fresh)
                cache['weather_data'] = processed
                cache['weather_timestamp'] = time.time()
                file_cache.set('weather_all_cities', processed)
                print(f"[Startup] Preloaded {len(processed)} cities in {time.time()-t0:.1f}s")
            else:
                print("[Startup] Weather preload returned no data")
        except Exception as e:
            print(f"[Startup] Preload error: {e}")
        finally:
            _weather_fetch_lock.release()

    threading.Thread(target=_startup_preload, daemon=True).start()


if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=Config.PORT,
        debug=(Config.FLASK_ENV == 'development')
    )
